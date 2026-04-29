"""
Interactive Review Analysis Dashboard — Keepa API Data
Streamlit app for exploring Amazon FBA business acquisition data.

Run: streamlit run dashboard.py
"""

import json
import os
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import streamlit as st
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta

import openpyxl
from openpyxl.utils import column_index_from_string

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

BIZ_COLORS = {
    "Wallaroo Wallets": {"primary": "#2c3e50", "accent": "#3498db"},
    "TeacherFav": {"primary": "#c0392b", "accent": "#e74c3c"},
}

PRODUCT_COLORS = [
    "#00d4aa", "#ff6b6b", "#4ecdc4", "#ffd93d", "#c084fc",
    "#ff8a5c", "#5ce1e6", "#ff5757", "#38b6ff", "#7bed9f",
]


# ─── Data Loading ───────────────────────────────────────────────────────────

@st.cache_data(ttl=60)
def load_data():
    with open(os.path.join(BASE_DIR, "keepa_deduped_catalog.json")) as f:
        catalog = json.load(f)

    biz_data = {}
    for biz_name in ["Wallaroo Wallets", "TeacherFav"]:
        products = sorted(
            [p for p in catalog if p["business"] == biz_name],
            key=lambda p: -p["shared_reviews"],
        )
        biz_data[biz_name] = {
            "products": products,
            "metrics": compute_metrics(products),
        }
    return biz_data


@st.cache_data
def load_pnl(biz_name):
    """Load P&L data from Excel if available. Returns dict of monthly data or None."""
    pnl_files = {
        "TeacherFav": os.path.join(BASE_DIR, "TeacherFav", "92221 - TeacherFav - P&L.xlsx"),
    }
    path = pnl_files.get(biz_name)
    if not path or not os.path.exists(path):
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Summary"]

    # Read month headers from row 3
    months = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if isinstance(val, datetime):
            months.append((col, val))

    key_rows = {
        "Total Revenue": 42,
        "Total Expenses": 133,
        "COGS": 95,
        "Gross Profit": 99,
        "Net Income": 145,
    }

    data = {}
    for label, row_num in key_rows.items():
        series = {}
        for col, month_dt in months:
            val = ws.cell(row=row_num, column=col).value
            if isinstance(val, (int, float)) and val != 0:
                series[month_dt.strftime("%Y-%m")] = round(val, 2)
        data[label] = series

    # Also grab margin rows (as percentages)
    margin_rows = {
        "Net Profit Margin (12mo trailing)": 150,
        "TaCOS (Total Ad Cost of Sales)": 149,
        "COGS %": 148,
    }
    for label, row_num in margin_rows.items():
        series = {}
        for col, month_dt in months:
            val = ws.cell(row=row_num, column=col).value
            if isinstance(val, (int, float)) and val != 0:
                series[month_dt.strftime("%Y-%m")] = round(val * 100, 1)
        data[label] = series

    return data


def compute_integrity_signals(products):
    """Analyze review/rating data for signs of manipulation or organic growth."""
    signals = []
    for p in products:
        ch = p["review_count_history"]
        rh = p["rating_history"]
        name = product_name(p, short=True)
        s = {"name": name, "asin": p["representative_asin"]}

        # 1. Review count drops (Amazon purges)
        drops = []
        for i in range(1, len(ch)):
            diff = ch[i]["count"] - ch[i - 1]["count"]
            if diff < -5:  # ignore tiny fluctuations
                drops.append({
                    "date": ch[i]["date"][:7],
                    "removed": abs(diff),
                    "from_count": ch[i - 1]["count"],
                })
        s["review_drops"] = drops
        s["total_removed"] = sum(d["removed"] for d in drops)

        # 2. Growth consistency — split into quarters, check for spikes
        if len(ch) >= 8:
            quarter_len = len(ch) // 4
            quarter_growths = []
            for q in range(4):
                start_idx = q * quarter_len
                end_idx = (q + 1) * quarter_len if q < 3 else len(ch) - 1
                q_months = max(1, (datetime.fromisoformat(ch[end_idx]["date"]) - datetime.fromisoformat(ch[start_idx]["date"])).days / 30.44)
                q_growth = (ch[end_idx]["count"] - ch[start_idx]["count"]) / q_months
                quarter_growths.append(round(q_growth, 1))
            s["quarter_growths"] = quarter_growths
            avg_growth = sum(quarter_growths) / len(quarter_growths) if quarter_growths else 0
            max_growth = max(quarter_growths) if quarter_growths else 0
            s["growth_spike"] = max_growth > avg_growth * 2 and avg_growth > 0
        else:
            s["quarter_growths"] = []
            s["growth_spike"] = False

        # 3. Rating stability — standard deviation
        if len(rh) >= 4:
            ratings = [r["rating"] for r in rh]
            mean_r = sum(ratings) / len(ratings)
            variance = sum((r - mean_r) ** 2 for r in ratings) / len(ratings)
            s["rating_std"] = round(variance ** 0.5, 3)
            s["rating_stable"] = s["rating_std"] < 0.3
        else:
            s["rating_std"] = None
            s["rating_stable"] = None

        # 4. Overall verdict
        red_flags = []
        if s["total_removed"] > 100:
            red_flags.append(f"Amazon removed {s['total_removed']:,} reviews ({len(drops)} events)")
        elif s["total_removed"] > 0:
            red_flags.append(f"Minor review removals: {s['total_removed']} reviews ({len(drops)} events)")
        if s["growth_spike"]:
            red_flags.append(f"Uneven growth pattern — quarterly rates: {s['quarter_growths']} reviews/mo")
        s["red_flags"] = red_flags
        s["looks_organic"] = len(red_flags) == 0 or (s["total_removed"] <= 20 and not s["growth_spike"])

        signals.append(s)
    return signals


def compute_metrics(products):
    m = {}
    m["n_products"] = len(products)
    m["n_variations"] = sum(p["num_variations"] for p in products)

    main = products[0]  # already sorted by reviews desc
    m["main_product"] = main
    m["main_asin"] = main["representative_asin"]
    m["main_title"] = main["title"] or ""
    # Use review count from filtered history if available, else fall back to static field
    main_ch = main["review_count_history"]
    m["main_reviews"] = main_ch[-1]["count"] if main_ch else main["shared_reviews"]
    m["main_rating"] = main["rating_history"][-1]["rating"] if main["rating_history"] else None

    # Weighted avg rating (use latest rating and review count within the filtered range)
    total_w, weighted_r = 0, 0
    for p in products:
        if p["rating_history"] and p["review_count_history"]:
            r = p["rating_history"][-1]["rating"]
            w = p["review_count_history"][-1]["count"]
            weighted_r += r * w
            total_w += w
    m["weighted_avg_rating"] = weighted_r / total_w if total_w > 0 else None

    # Total reviews = sum of latest review counts in filtered range
    m["total_reviews"] = sum(
        p["review_count_history"][-1]["count"] if p["review_count_history"] else p["shared_reviews"]
        for p in products
    )

    # Rating trend for main product
    rh = main["rating_history"]
    if len(rh) >= 4:
        mid = len(rh) // 2
        early = [r["rating"] for r in rh[:mid]]
        recent = [r["rating"] for r in rh[mid:]]
        m["main_early_r"] = sum(early) / len(early)
        m["main_recent_r"] = sum(recent) / len(recent)
        m["main_r_delta"] = m["main_recent_r"] - m["main_early_r"]
        m["main_r_dir"] = "IMPROVING" if m["main_r_delta"] > 0.05 else "DECLINING" if m["main_r_delta"] < -0.05 else "STABLE"
    else:
        m["main_early_r"] = m["main_recent_r"] = m["main_r_delta"] = 0
        m["main_r_dir"] = "INSUFFICIENT DATA"

    # Review growth
    ch = main["review_count_history"]
    if len(ch) >= 2:
        first_date = datetime.fromisoformat(ch[0]["date"])
        last_date = datetime.fromisoformat(ch[-1]["date"])
        months = max(1, (last_date - first_date).days / 30.44)
        m["main_growth_per_mo"] = (ch[-1]["count"] - ch[0]["count"]) / months
        m["main_first_date"] = ch[0]["date"][:10]
        m["main_last_date"] = ch[-1]["date"][:10]

        mid_idx = len(ch) // 2
        mid_date = datetime.fromisoformat(ch[mid_idx]["date"])
        early_months = max(1, (mid_date - first_date).days / 30.44)
        recent_months = max(1, (last_date - mid_date).days / 30.44)
        m["early_growth_mo"] = (ch[mid_idx]["count"] - ch[0]["count"]) / early_months
        m["recent_growth_mo"] = (ch[-1]["count"] - ch[mid_idx]["count"]) / recent_months
        m["growth_dir"] = "ACCELERATING" if m["recent_growth_mo"] > m["early_growth_mo"] * 1.1 else "DECELERATING" if m["recent_growth_mo"] < m["early_growth_mo"] * 0.9 else "STEADY"
    else:
        m["main_growth_per_mo"] = 0
        m["main_first_date"] = m["main_last_date"] = "?"
        m["early_growth_mo"] = m["recent_growth_mo"] = 0
        m["growth_dir"] = "UNKNOWN"

    # Rating and review trends over specific time windows (6mo, 1yr, 2yr)
    m["rating_periods"] = {}
    m["review_periods"] = {}
    if rh and len(rh) >= 2:
        latest_r_dt = datetime.fromisoformat(rh[-1]["date"])
        latest_rating = rh[-1]["rating"]
        for label, months_back in [("6mo", 6), ("1yr", 12), ("2yr", 24)]:
            cutoff_dt = latest_r_dt - timedelta(days=months_back * 30.44)
            cutoff_iso = cutoff_dt.isoformat()
            past_entries = [r for r in rh if r["date"] <= cutoff_iso]
            if past_entries:
                ref_entry = past_entries[-1]
                ref_rating = ref_entry["rating"]
                delta = latest_rating - ref_rating
                m["rating_periods"][label] = {
                    "from_month": ref_entry["date"][:7],
                    "from_rating": ref_rating,
                    "to_rating": latest_rating,
                    "delta": delta,
                }
    if ch and len(ch) >= 2:
        latest_c_dt = datetime.fromisoformat(ch[-1]["date"])
        latest_count = ch[-1]["count"]
        for label, months_back in [("6mo", 6), ("1yr", 12), ("2yr", 24)]:
            cutoff_dt = latest_c_dt - timedelta(days=months_back * 30.44)
            cutoff_iso = cutoff_dt.isoformat()
            past_entries = [c for c in ch if c["date"] <= cutoff_iso]
            if past_entries:
                ref_entry = past_entries[-1]
                ref_count = ref_entry["count"]
                added = latest_count - ref_count
                span_months = max(1, (latest_c_dt - datetime.fromisoformat(ref_entry["date"])).days / 30.44)
                m["review_periods"][label] = {
                    "from_month": ref_entry["date"][:7],
                    "from_count": ref_count,
                    "to_count": latest_count,
                    "added": added,
                    "per_month": added / span_months,
                }

    # Sales rank for main product — compute over specific time windows
    rank_data = main.get("monthly_avg_rank", {})
    if rank_data and len(rank_data) >= 2:
        rank_months = sorted(rank_data.keys())
        m["main_rank_recent"] = rank_data[rank_months[-1]]
        m["main_rank_recent_month"] = rank_months[-1]
        m["main_rank_best"] = min(rank_data.values())
        m["main_rank_worst"] = max(rank_data.values())

        # Compute rank change over 6mo, 1yr, 2yr windows
        latest_dt = datetime.strptime(rank_months[-1], "%Y-%m")
        m["main_rank_periods"] = {}
        for label, months_back in [("6mo", 6), ("1yr", 12), ("2yr", 24)]:
            cutoff_dt = latest_dt - timedelta(days=months_back * 30.44)
            cutoff_ym = cutoff_dt.strftime("%Y-%m")
            # Find the closest month at or after the cutoff
            past_months = [mo for mo in rank_months if mo <= cutoff_ym]
            if past_months:
                ref_month = past_months[-1]
                ref_rank = rank_data[ref_month]
                current_rank = m["main_rank_recent"]
                pct_change = (current_rank - ref_rank) / ref_rank * 100 if ref_rank > 0 else 0
                m["main_rank_periods"][label] = {
                    "from_month": ref_month,
                    "from_rank": ref_rank,
                    "to_rank": current_rank,
                    "pct_change": pct_change,
                }
    else:
        m["main_rank_recent"] = m["main_rank_best"] = m["main_rank_worst"] = None
        m["main_rank_recent_month"] = None
        m["main_rank_periods"] = {}

    return m


def get_date_range(biz_data):
    """Find the min/max dates across all products in all businesses."""
    all_dates = []
    for biz_name in biz_data:
        for p in biz_data[biz_name]["products"]:
            for r in p["rating_history"]:
                all_dates.append(r["date"][:10])
            for c in p["review_count_history"]:
                all_dates.append(c["date"][:10])
    if not all_dates:
        return date(2020, 1, 1), date.today()
    return (
        datetime.fromisoformat(min(all_dates)).date(),
        datetime.fromisoformat(max(all_dates)).date(),
    )


def filter_products(products, start_date, end_date):
    """Return a deep-copy of products with time-series data filtered to [start, end]."""
    start_str = start_date.isoformat()
    end_str = end_date.isoformat()
    filtered = []
    for p in products:
        fp = dict(p)
        fp["rating_history"] = [
            r for r in p["rating_history"] if start_str <= r["date"][:10] <= end_str
        ]
        fp["review_count_history"] = [
            c for c in p["review_count_history"] if start_str <= c["date"][:10] <= end_str
        ]
        fp["amazon_price_history"] = [
            pt for pt in p.get("amazon_price_history", []) if start_str <= pt["date"][:10] <= end_str
        ]
        fp["marketplace_price_history"] = [
            pt for pt in p.get("marketplace_price_history", []) if start_str <= pt["date"][:10] <= end_str
        ]
        for dict_field in ["monthly_avg_rank", "monthly_seller_count"]:
            if p.get(dict_field):
                fp[dict_field] = {
                    m: v for m, v in p[dict_field].items()
                    if start_str[:7] <= m <= end_str[:7]
                }
        fp["monthly_sold_history"] = [
            pt for pt in p.get("monthly_sold_history", []) if start_str <= pt["date"][:10] <= end_str
        ]
        fp["list_price_history"] = [
            pt for pt in p.get("list_price_history", []) if start_str <= pt["date"][:10] <= end_str
        ]
        fp["lightning_deals"] = [
            d for d in p.get("lightning_deals", []) if start_str <= d["date"][:10] <= end_str
        ]
        fp["stockouts"] = [
            s for s in p.get("stockouts", []) if start_str <= s["start"][:10] <= end_str
        ]
        fp["buybox_sellers"] = p.get("buybox_sellers", [])
        fp["buybox_changes"] = p.get("buybox_changes", 0)
        filtered.append(fp)
    return filtered


def amazon_link(asin):
    return f"https://www.amazon.com/dp/{asin}"


def keepa_link(asin):
    return f"https://keepa.com/#!product/1-{asin}"


# Short, distinctive names for sidebar and compact displays
PRODUCT_SHORT_NAMES = {
    "B01LYPPPUU": "Phone Wallet (main)",
    "B0BQ3RCZMQ": "MagSafe Wallet",
    "B07YS5CZYB": "Card Holder",
    "B0C2W24JY8": "Wristlet",
    "B08L879DLS": "Sand Timer 3-Pack (main)",
    "B09GF1VSBH": "Toothbrush Timer 2-Pack",
    "B09G6HHP2X": "Sand Timer 6-Pack",
    "B09Q9CBWNB": "Sand Timer 6-Pack Colorful",
    "B09VR2W9TV": "Hourglass 60min Large",
    "B09GP2HDCF": "Sand Timer 60min Black",
}


def product_name(product, short=False):
    """Return product display name. short=True uses compact sidebar names."""
    asin = product["representative_asin"]
    if short:
        return PRODUCT_SHORT_NAMES.get(asin, product["title"] or "Unknown")
    return product["title"] or "Unknown"


def short_title(title, max_len=80):
    if not title:
        return "Unknown"
    return title[:max_len] + "..." if len(title) > max_len else title


# ─── Chart Helpers ──────────────────────────────────────────────────────────

def detect_rating_spikes(rh, threshold=0.2):
    """Detect significant rating jumps/drops between consecutive data points."""
    spikes = []
    for i in range(1, len(rh)):
        delta = rh[i]["rating"] - rh[i - 1]["rating"]
        if abs(delta) >= threshold:
            spikes.append({
                "date": rh[i]["date"],
                "rating": rh[i]["rating"],
                "delta": delta,
                "direction": "drop" if delta < 0 else "jump",
            })
    return spikes


def detect_review_spikes(ch, threshold_multiplier=2.5):
    """Detect review purges (count drops) and unusual surges."""
    spikes = []
    if len(ch) < 3:
        return spikes

    # Calculate per-interval changes
    changes = []
    for i in range(1, len(ch)):
        days = max(1, (datetime.fromisoformat(ch[i]["date"]) - datetime.fromisoformat(ch[i - 1]["date"])).days)
        change = ch[i]["count"] - ch[i - 1]["count"]
        rate = change / (days / 30.44)  # normalize to per-month
        changes.append({"idx": i, "change": change, "rate": rate, "days": days})

    # Purges (any drop > 5)
    for c in changes:
        if c["change"] < -5:
            spikes.append({
                "date": ch[c["idx"]]["date"],
                "count": ch[c["idx"]]["count"],
                "change": c["change"],
                "direction": "purge",
            })

    # Surges (rate > threshold_multiplier * median positive rate)
    positive_rates = [c["rate"] for c in changes if c["rate"] > 0]
    if len(positive_rates) >= 4:
        positive_rates_sorted = sorted(positive_rates)
        median_rate = positive_rates_sorted[len(positive_rates_sorted) // 2]
        for c in changes:
            if c["rate"] > median_rate * threshold_multiplier and c["rate"] > 20:
                spikes.append({
                    "date": ch[c["idx"]]["date"],
                    "count": ch[c["idx"]]["count"],
                    "change": c["change"],
                    "direction": "surge",
                })

    return spikes


def render_notable_events(target_products, label):
    """Render the two side-by-side notable events tables for a list of products."""
    st.subheader(f"Notable Events ({label})")
    st.caption("Click column headers to sort — sort by Change to see the largest spikes first.")

    all_rating_events = []
    all_review_events = []
    for p in target_products:
        name = product_name(p, short=True)
        for s in detect_rating_spikes(p["rating_history"]):
            all_rating_events.append({
                "Product": name,
                "Date": s["date"][:7],
                "Type": "Drop" if s["direction"] == "drop" else "Jump",
                "From": round(s["rating"] - s["delta"], 1),
                "To": round(s["rating"], 1),
                "Change": round(s["delta"], 2),
                "Significance": "Possible quality issue or product change" if s["direction"] == "drop" else "Recovery or data correction",
            })
        for s in detect_review_spikes(p["review_count_history"]):
            all_review_events.append({
                "Product": name,
                "Date": s["date"][:7],
                "Type": "Purge" if s["direction"] == "purge" else "Surge",
                "Change": s["change"],
                "Count After": s["count"],
                "Significance": "Amazon removed reviews — past review manipulation flagged" if s["direction"] == "purge" else "Unusually high number of new reviews in a short period — worth checking if reviews are genuine",
            })

    all_price_events = []
    for p in target_products:
        name = product_name(p, short=True)
        for s in detect_price_spikes(p.get("marketplace_price_history", [])):
            all_price_events.append({
                "Product": name,
                "Date": s["date"][:7],
                "Type": "Increase" if s["direction"] == "increase" else "Decrease",
                "From": f"${s['prev_price']:.2f}",
                "To": f"${s['price']:.2f}",
                "Change": f"{s['pct_change']:+.0%}",
                "Significance": "Price increase — higher margin per sale but may reduce the number of sales" if s["direction"] == "increase" else "Price drop — lower margin per sale, likely due to competition or clearing inventory",
            })

    col_r, col_v, col_p = st.columns(3)
    with col_r:
        st.markdown("**Rating Changes** (drops/jumps > 0.2)")
        if all_rating_events:
            df_r = pd.DataFrame(all_rating_events).sort_values("Change")
            st.dataframe(df_r, use_container_width=True, hide_index=True)
        else:
            st.success("No significant rating spikes detected.")

    with col_v:
        st.markdown("**Review Changes** (purges & unusual surges)")
        if all_review_events:
            df_v = pd.DataFrame(all_review_events).sort_values("Change")
            st.dataframe(df_v, use_container_width=True, hide_index=True)
        else:
            st.success("No review purges or unusual surges detected.")

    with col_p:
        st.markdown("**Price Changes** (>20% between data points)")
        if all_price_events:
            df_p = pd.DataFrame(all_price_events).sort_values("Change")
            st.dataframe(df_p, use_container_width=True, hide_index=True)
        else:
            st.success("No significant price changes detected.")


def make_rating_chart(products, title, colors=None):
    fig = go.Figure()
    for i, p in enumerate(products):
        rh = p["rating_history"]
        if not rh:
            continue
        dates = [r["date"] for r in rh]
        ratings = [r["rating"] for r in rh]
        color = (colors or PRODUCT_COLORS)[i % len(colors or PRODUCT_COLORS)]
        fig.add_trace(go.Scatter(
            x=dates, y=ratings, mode="lines+markers",
            name=f"{product_name(p, short=True)} ({p['shared_reviews']:,} rev)",
            line=dict(width=2, color=color),
            marker=dict(size=4),
        ))

    fig.add_hline(y=4.0, line_dash="dot", line_color="green", opacity=0.3,
                  annotation_text="4.0", annotation_position="left")
    layout_kwargs = dict(
        title=dict(text=f"{title}<br><sup>Point-in-time snapshot — each dot is the product's star rating on that date. Source: Keepa API (keepa.com)</sup>"),
        yaxis_title="Rating",
        yaxis=dict(range=[2.5, 5.1], dtick=0.5),
        height=450,
        legend=dict(font=dict(size=10)),
        margin=dict(l=60, r=20, t=70, b=40),
    )
    if len(products) > 1:
        layout_kwargs["legend"] = dict(orientation="h", yanchor="top", y=-0.12, xanchor="center", x=0.5, font=dict(size=10))
        layout_kwargs["margin"] = dict(l=60, r=20, t=70, b=80)
        layout_kwargs["height"] = 480
    fig.update_layout(**layout_kwargs)
    return fig


def make_review_count_chart(products, title, colors=None):
    fig = go.Figure()
    for i, p in enumerate(products):
        ch = p["review_count_history"]
        if not ch:
            continue
        dates = [c["date"] for c in ch]
        counts = [c["count"] for c in ch]
        color = (colors or PRODUCT_COLORS)[i % len(colors or PRODUCT_COLORS)]
        fig.add_trace(go.Scatter(
            x=dates, y=counts, mode="lines",
            name=f"{product_name(p, short=True)} ({counts[-1]:,} current)",
            line=dict(width=2.5, color=color),
        ))

    layout_kwargs = dict(
        title=dict(
            text=f"{title}<br><sup>Cumulative running total — shows all reviews to date, not new reviews per period. Drops indicate reviews removed by Amazon. Source: Keepa API (keepa.com)</sup>",
        ),
        yaxis_title="Cumulative Review Count",
        height=450,
        yaxis=dict(rangemode="tozero"),
        legend=dict(font=dict(size=10)),
        margin=dict(l=60, r=20, t=70, b=40),
    )
    if len(products) > 1:
        layout_kwargs["legend"] = dict(orientation="h", yanchor="top", y=-0.12, xanchor="center", x=0.5, font=dict(size=10))
        layout_kwargs["margin"] = dict(l=60, r=20, t=70, b=80)
        layout_kwargs["height"] = 480
    fig.update_layout(**layout_kwargs)
    return fig


def make_combined_chart(products, metric_type, title, colors=None):
    """Create a combined (summed) chart across all products."""
    # Collect all dates and sum values
    if metric_type == "reviews":
        all_series = []
        for p in products:
            ch = p["review_count_history"]
            if ch:
                all_series.append(pd.Series(
                    {c["date"][:10]: c["count"] for c in ch}
                ))
        if not all_series:
            return go.Figure()

        # Resample to monthly, forward-fill, then sum
        combined = pd.DataFrame(all_series).T
        combined.index = pd.to_datetime(combined.index)
        combined = combined.sort_index().ffill().fillna(0)
        monthly = combined.resample("MS").last().ffill()
        totals = monthly.sum(axis=1)

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=totals.index, y=totals.values, mode="lines",
            name="All Products Combined",
            line=dict(width=3, color="#00d4aa"),
            fill="tozeroy", fillcolor="rgba(0,212,170,0.15)",
        ))
        fig.update_layout(
            title=dict(text=f"{title}<br><sup>Source: Keepa API (keepa.com)</sup>"),
            yaxis_title="Total Review Count (All Products)",
            height=400,
            yaxis=dict(rangemode="tozero"),
            margin=dict(l=60, r=20, t=50, b=40),
        )
        return fig

    elif metric_type == "rating":
        # Weighted average rating over time
        all_dates = set()
        for p in products:
            for r in p["rating_history"]:
                all_dates.add(r["date"][:10])
        all_dates = sorted(all_dates)

        if not all_dates:
            return go.Figure()

        # Build per-product rating series, forward-fill
        product_ratings = {}
        product_weights = {}
        for p in products:
            rh = {r["date"][:10]: r["rating"] for r in p["rating_history"]}
            ch = {c["date"][:10]: c["count"] for c in p["review_count_history"]}
            product_ratings[p["representative_asin"]] = rh
            product_weights[p["representative_asin"]] = ch

        dates_out = []
        ratings_out = []

        last_ratings = {asin: None for asin in product_ratings}
        last_weights = {asin: 1 for asin in product_ratings}

        for d in all_dates:
            for asin in product_ratings:
                if d in product_ratings[asin]:
                    last_ratings[asin] = product_ratings[asin][d]
                if d in product_weights.get(asin, {}):
                    last_weights[asin] = product_weights[asin][d]

            total_w = 0
            weighted_sum = 0
            for asin in product_ratings:
                if last_ratings[asin] is not None:
                    w = last_weights[asin]
                    weighted_sum += last_ratings[asin] * w
                    total_w += w

            if total_w > 0:
                dates_out.append(d)
                ratings_out.append(weighted_sum / total_w)

        # Downsample to monthly for readability
        df = pd.DataFrame({"date": pd.to_datetime(dates_out), "rating": ratings_out})
        df = df.set_index("date").resample("MS").mean().dropna()

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=df.index, y=df["rating"], mode="lines+markers",
            name="Weighted Avg Rating (All Products)",
            line=dict(width=3, color="#00d4aa"),
            marker=dict(size=5),
        ))
        fig.add_hline(y=4.0, line_dash="dot", line_color="green", opacity=0.3)
        fig.update_layout(
            title=dict(text=f"{title}<br><sup>Source: Keepa API (keepa.com)</sup>"),
            yaxis_title="Weighted Avg Rating",
            yaxis=dict(range=[2.5, 5.1], dtick=0.5),
            height=400,
            margin=dict(l=60, r=20, t=50, b=40),
        )
        return fig


def make_price_chart(products, title, colors=None):
    """Create a price history chart showing marketplace (seller) price over time."""
    fig = go.Figure()
    has_data = False
    for i, p in enumerate(products):
        ph = p.get("marketplace_price_history", [])
        if not ph:
            continue
        has_data = True
        dates = [pt["date"] for pt in ph]
        prices = [pt["price"] for pt in ph]
        color = (colors or PRODUCT_COLORS)[i % len(colors or PRODUCT_COLORS)]
        latest_price = prices[-1] if prices else 0
        fig.add_trace(go.Scatter(
            x=dates, y=prices, mode="lines+markers",
            name=f"{product_name(p, short=True)} (${latest_price:.2f})",
            line=dict(width=2, color=color),
            marker=dict(size=4),
        ))

    if not has_data:
        fig.add_annotation(text="No price data available", xref="paper", yref="paper",
                           x=0.5, y=0.5, showarrow=False, font=dict(size=16))

    layout_kwargs = dict(
        title=dict(text=f"{title}<br><sup>Marketplace (seller) price over time — the price customers pay on Amazon. Source: Keepa API (keepa.com)</sup>"),
        yaxis_title="Price (USD)",
        yaxis=dict(tickprefix="$", rangemode="tozero"),
        height=450,
        legend=dict(font=dict(size=10)),
        margin=dict(l=60, r=20, t=70, b=40),
    )
    if len(products) > 1:
        layout_kwargs["legend"] = dict(orientation="h", yanchor="top", y=-0.12, xanchor="center", x=0.5, font=dict(size=10))
        layout_kwargs["margin"] = dict(l=60, r=20, t=70, b=80)
        layout_kwargs["height"] = 480
    fig.update_layout(**layout_kwargs)
    return fig


def _make_price_vs_velocity_chart(product, title):
    """Dual-axis chart: price over time vs review velocity."""
    ph = product.get("marketplace_price_history", [])
    ch = product.get("review_count_history", [])
    if not ph or len(ch) < 4:
        return None

    price_by_month = {}
    for pt in ph:
        price_by_month.setdefault(pt["date"][:7], []).append(pt["price"])
    monthly_price = {m: sum(v) / len(v) for m, v in sorted(price_by_month.items())}

    count_by_month = {}
    for c in ch:
        count_by_month[c["date"][:7]] = c["count"]
    ms = sorted(count_by_month.keys())
    vel_by_month = {}
    for i in range(1, len(ms)):
        vel_by_month[ms[i]] = max(0, count_by_month[ms[i]] - count_by_month[ms[i - 1]])

    common = sorted(set(monthly_price) & set(vel_by_month))
    if len(common) < 3:
        return None

    dates = [datetime.strptime(m, "%Y-%m") for m in common]
    prices = [monthly_price[m] for m in common]
    vels = [vel_by_month[m] for m in common]

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Scatter(x=dates, y=prices, mode="lines+markers", name="Avg Price",
                   line=dict(width=2.5, color="#ffd93d"), marker=dict(size=4)),
        secondary_y=False,
    )
    fig.add_trace(
        go.Bar(x=dates, y=vels, name="New Reviews/mo",
               marker_color="rgba(0,212,170,0.4)"),
        secondary_y=True,
    )
    fig.update_layout(
        title=dict(text=f"{title}<br><sup>Gold line = seller price. Green bars = new reviews that month (purge months excluded).</sup>"),
        yaxis=dict(title="Price (USD)", tickprefix="$"),
        yaxis2=dict(title="New Reviews / Month", rangemode="tozero"),
        height=420,
        legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5, font=dict(size=10)),
        margin=dict(l=60, r=60, t=70, b=80),
    )
    return fig


def _make_price_vs_rating_chart(product, title):
    """Dual-axis chart: price over time vs rating."""
    ph = product.get("marketplace_price_history", [])
    rh = product.get("rating_history", [])
    if not ph or len(rh) < 4:
        return None

    price_by_month = {}
    for pt in ph:
        price_by_month.setdefault(pt["date"][:7], []).append(pt["price"])
    monthly_price = {m: sum(v) / len(v) for m, v in sorted(price_by_month.items())}

    rating_by_month = {}
    for r in rh:
        rating_by_month.setdefault(r["date"][:7], []).append(r["rating"])
    monthly_rating = {m: sum(v) / len(v) for m, v in sorted(rating_by_month.items())}

    common = sorted(set(monthly_price) & set(monthly_rating))
    if len(common) < 3:
        return None

    dates = [datetime.strptime(m, "%Y-%m") for m in common]
    prices = [monthly_price[m] for m in common]
    ratings = [monthly_rating[m] for m in common]

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Scatter(x=dates, y=prices, mode="lines+markers", name="Avg Price",
                   line=dict(width=2.5, color="#ffd93d"), marker=dict(size=4)),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(x=dates, y=ratings, mode="lines+markers", name="Rating",
                   line=dict(width=2.5, color="#c084fc"), marker=dict(size=5)),
        secondary_y=True,
    )
    fig.update_layout(
        title=dict(text=f"{title}<br><sup>Gold line = seller price. Purple line = star rating. Look for inverse patterns (price up, rating down).</sup>"),
        yaxis=dict(title="Price (USD)", tickprefix="$"),
        yaxis2=dict(title="Rating", range=[2.5, 5.1], dtick=0.5),
        height=420,
        legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5, font=dict(size=10)),
        margin=dict(l=60, r=60, t=70, b=80),
    )
    return fig


def _make_price_scatter(product):
    """Scatter plots with best-fit regression: price vs rating and price vs velocity."""
    ph = product.get("marketplace_price_history", [])
    ch = product.get("review_count_history", [])
    rh = product.get("rating_history", [])
    if not ph:
        return None, None

    # Build monthly aggregates
    price_by_m = {}
    for pt in ph:
        price_by_m.setdefault(pt["date"][:7], []).append(pt["price"])
    monthly_price = {m: sum(v) / len(v) for m, v in price_by_m.items()}

    def best_fit(x, y):
        """Try poly degrees 1-3, pick best by adjusted R². Return (degree, coeffs, r2)."""
        n = len(x)
        best_deg, best_r2, best_coeffs = 1, -1, None
        for deg in range(1, min(4, n - 1)):
            coeffs = np.polyfit(x, y, deg)
            y_pred = np.polyval(coeffs, x)
            ss_res = np.sum((y - y_pred) ** 2)
            ss_tot = np.sum((y - np.mean(y)) ** 2)
            if ss_tot == 0:
                continue
            r2 = 1 - ss_res / ss_tot
            # Adjusted R² penalizes higher degrees
            adj_r2 = 1 - (1 - r2) * (n - 1) / max(1, n - deg - 1)
            if adj_r2 > best_r2:
                best_deg, best_r2, best_coeffs = deg, adj_r2, coeffs
        return best_deg, best_coeffs, best_r2

    def make_scatter_fig(prices, values, y_label, title_suffix):
        x = np.array(prices)
        y = np.array(values)
        deg, coeffs, r2 = best_fit(x, y)

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=x, y=y, mode="markers", name="Monthly data",
            marker=dict(size=8, color="#00d4aa", opacity=0.7),
        ))

        # Smooth fit line
        x_fit = np.linspace(x.min(), x.max(), 100)
        y_fit = np.polyval(coeffs, x_fit)
        deg_labels = {1: "Linear", 2: "Quadratic", 3: "Cubic"}
        fig.add_trace(go.Scatter(
            x=x_fit, y=y_fit, mode="lines",
            name=f"{deg_labels.get(deg, f'Poly-{deg}')} fit (R²={r2:.2f})",
            line=dict(width=2.5, color="#ff6b6b", dash="dash"),
        ))

        fig.update_layout(
            title=dict(text=f"Price vs {title_suffix}<br><sup>{deg_labels.get(deg, f'Poly-{deg}')} best fit — R² = {r2:.2f} {'(good fit)' if r2 >= 0.4 else '(weak fit — no clear pattern)'}</sup>"),
            xaxis=dict(title="Price (USD)", tickprefix="$"),
            yaxis=dict(title=y_label),
            height=420,
            legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5, font=dict(size=10)),
            margin=dict(l=60, r=20, t=70, b=80),
        )
        return fig

    # Price vs Rating scatter
    fig_rating = None
    rating_by_m = {}
    for r in rh:
        rating_by_m.setdefault(r["date"][:7], []).append(r["rating"])
    monthly_rating = {m: sum(v) / len(v) for m, v in rating_by_m.items()}
    common_r = sorted(set(monthly_price) & set(monthly_rating))
    if len(common_r) >= 5:
        fig_rating = make_scatter_fig(
            [monthly_price[m] for m in common_r],
            [monthly_rating[m] for m in common_r],
            "Rating", "Rating",
        )

    # Price vs Velocity scatter
    fig_velocity = None
    count_by_m = {}
    for c in ch:
        count_by_m[c["date"][:7]] = c["count"]
    ms = sorted(count_by_m.keys())
    vel_by_m = {}
    for i in range(1, len(ms)):
        vel_by_m[ms[i]] = max(0, count_by_m[ms[i]] - count_by_m[ms[i - 1]])
    common_v = sorted(set(monthly_price) & set(vel_by_m))
    if len(common_v) >= 5:
        fig_velocity = make_scatter_fig(
            [monthly_price[m] for m in common_v],
            [vel_by_m[m] for m in common_v],
            "New Reviews / Month", "Review Velocity",
        )

    return fig_rating, fig_velocity


def _ai_caption(text):
    """Render a standardized AI analysis caption."""
    st.caption(f"*AI analysis (Claude):* {text}")


def _analyze_rating(products):
    """Generate AI analysis for a rating chart."""
    summaries = []
    for p in products:
        rh = p["rating_history"]
        if len(rh) < 2:
            continue
        name = product_name(p, short=True)
        first_r, last_r = rh[0]["rating"], rh[-1]["rating"]
        delta = last_r - first_r
        if abs(delta) <= 0.1:
            summaries.append(f"{name} stable at {last_r:.1f}")
        elif delta > 0:
            summaries.append(f"{name} up from {first_r:.1f} to {last_r:.1f}")
        else:
            summaries.append(f"{name} down from {first_r:.1f} to {last_r:.1f}")
    if not summaries:
        return None
    # Factual stake
    below_threshold = [p for p in products if p["rating_history"] and p["rating_history"][-1]["rating"] < 4.3]
    if below_threshold:
        names = ", ".join(product_name(p, short=True) for p in below_threshold)
        stake = f" Products below 4.3 ({names}) are rated lower than most competitors — fewer customers will choose them over alternatives."
    else:
        stake = " All products above 4.3 — customers generally trust these ratings, which helps maintain sales."
    return "; ".join(summaries) + "." + stake


def _analyze_review_count(products):
    """Generate AI analysis for a review count chart."""
    summaries = []
    total_purges = 0
    for p in products:
        ch = p["review_count_history"]
        if len(ch) < 2:
            continue
        name = product_name(p, short=True)
        first_c, last_c = ch[0]["count"], ch[-1]["count"]
        months = max(1, (datetime.fromisoformat(ch[-1]["date"]) - datetime.fromisoformat(ch[0]["date"])).days / 30.44)
        rate = (last_c - first_c) / months
        drops = sum(1 for i in range(1, len(ch)) if ch[i]["count"] - ch[i-1]["count"] < -5)
        total_purges += drops
        drop_note = f", {drops} purge(s)" if drops else ""
        summaries.append(f"{name}: {last_c:,} reviews at ~{rate:.0f}/mo{drop_note}")
    if not summaries:
        return None
    if total_purges > 0:
        stake = f" Amazon purged reviews {total_purges} time(s) across the portfolio — each purge permanently removes reviews, making the listing less convincing to new shoppers."
    else:
        stake = " Zero Amazon purges detected — the review base is intact, which helps convince new customers to buy."
    return "; ".join(summaries) + "." + stake


def _analyze_velocity(velocity_values):
    """Generate AI analysis for a review velocity bar chart."""
    if not velocity_values:
        return None
    positive = [v for v in velocity_values if v > 0]
    negative = [v for v in velocity_values if v < 0]
    avg_pos = sum(positive) / len(positive) if positive else 0
    recent = velocity_values[-6:] if len(velocity_values) > 6 else velocity_values
    earlier = velocity_values[:-6] if len(velocity_values) > 6 else []
    avg_recent = sum(recent) / len(recent) if recent else 0
    avg_earlier = sum(earlier) / len(earlier) if earlier else avg_recent

    trend = "accelerating" if avg_recent > avg_earlier * 1.2 else "decelerating" if avg_recent < avg_earlier * 0.8 else "steady"
    purge_note = f" {len(negative)} months had net review removals by Amazon." if negative else ""
    if trend == "decelerating":
        stake = f" At {avg_recent:.0f}/mo (down from {avg_earlier:.0f}/mo), fewer new reviews are coming in — this usually means fewer people are buying the product."
    elif trend == "accelerating":
        stake = f" At {avg_recent:.0f}/mo (up from {avg_earlier:.0f}/mo), more new reviews are coming in — this usually means more people are buying the product."
    else:
        stake = f" Steady at ~{avg_pos:.0f}/mo — the product is getting new reviews at a consistent rate."
    return f"The rate of new reviews is {trend}.{purge_note}{stake}"


def _analyze_price(product):
    """Generate AI analysis for a price history chart."""
    ph = product.get("marketplace_price_history", [])
    if not ph or len(ph) < 2:
        return None
    prices = [pt["price"] for pt in ph]
    first_p, last_p = prices[0], prices[-1]
    min_p, max_p = min(prices), max(prices)
    pct_change = (last_p - first_p) / first_p * 100 if first_p > 0 else 0
    volatility = (max_p - min_p) / min_p * 100 if min_p > 0 else 0

    if abs(pct_change) < 5:
        trend = f"Price stable at \\${last_p:.2f}"
    elif pct_change > 0:
        trend = f"Price up {pct_change:.0f}% from \\${first_p:.2f} to \\${last_p:.2f}"
    else:
        trend = f"Price down {abs(pct_change):.0f}% from \\${first_p:.2f} to \\${last_p:.2f}"

    if volatility > 50:
        stake = f" Range of \\${min_p:.2f}–\\${max_p:.2f} ({volatility:.0f}% spread) — the price has moved a lot, so past margins may not be a good guide for future margins."
    elif volatility > 20:
        stake = f" Range of \\${min_p:.2f}–\\${max_p:.2f} — some price variation, which means margins have shifted depending on the period."
    else:
        stake = f" Tight range of \\${min_p:.2f}–\\${max_p:.2f} — stable pricing, so margins should be predictable."
    return f"{trend}.{stake}"


def _analyze_sales_rank(products):
    """Generate AI analysis for a sales rank chart."""
    summaries = []
    for p in products:
        rank_data = p.get("monthly_avg_rank", {})
        if len(rank_data) < 2:
            continue
        name = product_name(p, short=True)
        months = sorted(rank_data.keys())
        recent_rank = rank_data[months[-1]]
        recent_month = months[-1]
        if recent_rank < 1000:
            tier = "top 1,000"
        elif recent_rank < 10000:
            tier = f"rank ~{recent_rank:,.0f} (top 10K)"
        elif recent_rank < 50000:
            tier = f"rank ~{recent_rank:,.0f} (top 50K)"
        else:
            tier = f"rank ~{recent_rank:,.0f}"
        # Show change over 1yr if available, else full range
        latest_dt = datetime.strptime(months[-1], "%Y-%m")
        cutoff_1yr = (latest_dt - timedelta(days=365)).strftime("%Y-%m")
        past_months_1yr = [mo for mo in months if mo <= cutoff_1yr]
        if past_months_1yr:
            ref_month = past_months_1yr[-1]
            ref_rank = rank_data[ref_month]
            pct = (recent_rank - ref_rank) / ref_rank * 100 if ref_rank > 0 else 0
            change_text = f"{pct:+.0f}% vs 1yr ago (~{ref_rank:,.0f} in {ref_month})"
        else:
            ref_month = months[0]
            ref_rank = rank_data[ref_month]
            pct = (recent_rank - ref_rank) / ref_rank * 100 if ref_rank > 0 else 0
            change_text = f"{pct:+.0f}% vs {ref_month} (~{ref_rank:,.0f})"
        summaries.append(f"{name}: {tier} as of {recent_month}, {change_text}")
    if not summaries:
        return None
    worsening = [s for s in summaries if "+%" in s or "worsening" in s.lower()]
    if any("+" in s.split("vs")[0] for s in summaries if "vs" in s):
        stake = " Higher rank number = fewer daily sales compared to other products in the same category."
    else:
        stake = " Lower rank number = more daily sales compared to others in the category."
    return "; ".join(summaries) + "." + stake


def _analyze_combined_rating(products):
    """Generate AI analysis for combined weighted rating chart."""
    if not products:
        return None
    total_w, weighted_sum = 0, 0
    for p in products:
        if p["rating_history"] and p["review_count_history"]:
            r = p["rating_history"][-1]["rating"]
            w = p["review_count_history"][-1]["count"]
            weighted_sum += r * w
            total_w += w
    if total_w == 0:
        return None
    avg = weighted_sum / total_w
    if avg >= 4.3:
        stake = f"At {avg:.2f}, the portfolio's average rating is strong — customers generally trust products rated above 4.3, which helps maintain steady sales."
    else:
        stake = f"At {avg:.2f}, the portfolio's average rating is below 4.3 — lower ratings make it harder to compete because customers tend to pick higher-rated alternatives."
    return stake


def _analyze_combined_reviews(products):
    """Generate AI analysis for combined review count chart."""
    total = sum(
        p["review_count_history"][-1]["count"] for p in products if p["review_count_history"]
    )
    n = sum(1 for p in products if p["review_count_history"])
    concentration = max(
        (p["review_count_history"][-1]["count"] / total * 100) for p in products if p["review_count_history"]
    ) if total > 0 else 0
    top_product = max(products, key=lambda p: p["review_count_history"][-1]["count"] if p["review_count_history"] else 0)
    top_name = product_name(top_product, short=True)
    stake = f"{total:,} total reviews across {n} products. {top_name} holds {concentration:.0f}% of all reviews — "
    if concentration > 70:
        stake += "the business depends heavily on this one product. If anything happens to it, the majority of the portfolio's reviews are at risk."
    elif concentration > 50:
        stake += "the portfolio has moderate concentration risk in one product."
    else:
        stake += "reviews are spread across products, reducing dependence on any single one."
    return stake


def _compute_price_insight(product):
    """Analyze price vs review velocity and price vs rating. Return insight text only if meaningful."""
    ph = product.get("marketplace_price_history", [])
    ch = product.get("review_count_history", [])
    rh = product.get("rating_history", [])
    if not ph:
        return None

    prices_all = [pt["price"] for pt in ph]
    price_min, price_max = min(prices_all), max(prices_all)

    # Monthly price
    price_by_month = {}
    for pt in ph:
        price_by_month.setdefault(pt["date"][:7], []).append(pt["price"])
    monthly_price = {m: sum(v) / len(v) for m, v in price_by_month.items()}

    findings = []

    # Price vs velocity (need 10+ months for reliability)
    if len(ch) >= 4:
        count_by_month = {}
        for c in ch:
            count_by_month[c["date"][:7]] = c["count"]
        ms = sorted(count_by_month.keys())
        vel_by_month = {}
        for i in range(1, len(ms)):
            vel_by_month[ms[i]] = max(0, count_by_month[ms[i]] - count_by_month[ms[i - 1]])
        common = sorted(set(monthly_price) & set(vel_by_month))
        if len(common) >= 10:
            pr = np.array([monthly_price[m] for m in common])
            ve = np.array([vel_by_month[m] for m in common])
            if pr.std() > 0 and ve.std() > 0:
                corr = float(np.corrcoef(pr, ve)[0, 1])
                if abs(corr) >= 0.4:
                    if corr > 0:
                        findings.append(f"Price increases correlate with more new reviews per month (r={corr:.2f} over {len(common)} months) — price increases didn't hurt sales. The product held up well at higher prices.")
                    else:
                        findings.append(f"Price increases correlate with fewer new reviews per month (r={corr:.2f} over {len(common)} months) — customers buy less when the price goes up. Higher margins from price increases are offset by fewer sales.")

    # Price vs rating (need 8+ months)
    if len(rh) >= 4:
        rating_by_month = {}
        for r in rh:
            rating_by_month.setdefault(r["date"][:7], []).append(r["rating"])
        monthly_rating = {m: sum(v) / len(v) for m, v in rating_by_month.items()}
        common = sorted(set(monthly_price) & set(monthly_rating))
        if len(common) >= 8:
            pr = np.array([monthly_price[m] for m in common])
            ra = np.array([monthly_rating[m] for m in common])
            if pr.std() > 0 and ra.std() > 0:
                corr = float(np.corrcoef(pr, ra)[0, 1])
                if abs(corr) >= 0.4:
                    if corr < 0:
                        findings.append(f"Higher prices correlate with lower ratings (r={corr:.2f} over {len(common)} months) — customers leave worse reviews when the price is higher, suggesting they expect more for the money.")
                    else:
                        findings.append(f"Higher prices correlate with higher ratings (r={corr:.2f} over {len(common)} months) — customers rate the product well even at higher prices, suggesting they see it as worth the cost.")

    # Always show the price range context
    header = f"**Pricing analysis** — Price ranged from \\${price_min:.2f} to \\${price_max:.2f} over {len(ph)} data points."

    if findings:
        return header + "\n\n" + "\n\n".join(findings)
    else:
        return header + " No significant relationship found between price changes and new reviews or ratings — price changes within this range haven't noticeably affected customer behavior, giving the operator flexibility to adjust prices."



def detect_price_spikes(ph, threshold=0.20):
    """Detect significant price changes (>threshold fraction) between consecutive data points."""
    spikes = []
    for i in range(1, len(ph)):
        prev_price = ph[i - 1]["price"]
        curr_price = ph[i]["price"]
        if prev_price > 0:
            pct_change = (curr_price - prev_price) / prev_price
            if abs(pct_change) >= threshold:
                spikes.append({
                    "date": ph[i]["date"],
                    "price": curr_price,
                    "prev_price": prev_price,
                    "pct_change": pct_change,
                    "direction": "increase" if pct_change > 0 else "decrease",
                })
    return spikes


def make_sales_rank_chart(products, title, zoomed=False):
    fig = go.Figure()
    all_ranks = []
    for i, p in enumerate(products):
        rank_data = p.get("monthly_avg_rank", {})
        if not rank_data or len(rank_data) < 2:
            continue
        months = sorted(rank_data.keys())
        dates = [datetime.strptime(m, "%Y-%m") for m in months]
        ranks = [rank_data[m] for m in months]
        all_ranks.extend(ranks)
        fig.add_trace(go.Scatter(
            x=dates, y=ranks, mode="lines",
            name=product_name(p, short=True),
            line=dict(width=1.8, color=PRODUCT_COLORS[i % len(PRODUCT_COLORS)]),
        ))

    if zoomed and all_ranks:
        # Auto-scale to data range with 10% padding
        data_min, data_max = min(all_ranks), max(all_ranks)
        padding = max((data_max - data_min) * 0.1, 100)
        y_lo = max(0, data_min - padding)
        y_hi = data_max + padding
        # Only add bands that overlap with the visible range
        rank_bands = [
            (0, 1_000, "#1–1K", "rgba(0,212,170,0.12)"),
            (1_000, 10_000, "#1K–10K", "rgba(56,182,255,0.10)"),
            (10_000, 50_000, "#10K–50K", "rgba(255,217,61,0.08)"),
            (50_000, 200_000, "#50K–200K", "rgba(255,107,107,0.08)"),
            (200_000, 1_000_000, "#200K–1M", "rgba(192,132,252,0.08)"),
        ]
        for band_y0, band_y1, label, color in rank_bands:
            if band_y1 > y_lo and band_y0 < y_hi:
                fig.add_hrect(y0=max(band_y0, y_lo), y1=min(band_y1, y_hi), fillcolor=color, line_width=0,
                              annotation_text=label, annotation_position="top left",
                              annotation=dict(font_size=10, font_color="rgba(255,255,255,0.5)"))
        yaxis_config = dict(range=[y_hi, y_lo])  # reversed: higher number at bottom
        subtitle = "Zoomed in to show trends and fluctuations. Same data as above, scaled to the actual range."
    else:
        # Full-range view with all bands
        rank_bands = [
            (0, 1_000, "#1–1K", "rgba(0,212,170,0.12)"),
            (1_000, 10_000, "#1K–10K", "rgba(56,182,255,0.10)"),
            (10_000, 50_000, "#10K–50K", "rgba(255,217,61,0.08)"),
            (50_000, 200_000, "#50K–200K", "rgba(255,107,107,0.08)"),
            (200_000, 1_000_000, "#200K–1M", "rgba(192,132,252,0.08)"),
        ]
        for y0, y1, label, color in rank_bands:
            fig.add_hrect(y0=y0, y1=y1, fillcolor=color, line_width=0)
        yaxis_config = dict(autorange="reversed")
        subtitle = "Sales rank shows how this product's sales compare to every other product in the same Amazon category.<br>Rank #1 sells the most. Rank 100,000 means 99,999 products sell more."

    fig.update_layout(
        title=dict(text=f"{title}<br><sup>{subtitle}</sup>"),
        yaxis_title="Sales Rank (lower = more sales)",
        yaxis=yaxis_config,
        height=450,
        legend=dict(orientation="h", yanchor="top", y=-0.12, xanchor="center", x=0.5, font=dict(size=10)),
        margin=dict(l=60, r=20, t=60, b=80),
    )
    return fig


# ─── Pages ──────────────────────────────────────────────────────────────────

def page_executive_summary(biz_data):
    st.header("Executive Summary")
    st.caption("Based on Keepa API data | All products deduplicated by parent ASIN")

    col1, col2 = st.columns(2)

    for col, biz_name in zip([col1, col2], ["Wallaroo Wallets", "TeacherFav"]):
        m = biz_data[biz_name]["metrics"]
        with col:
            st.subheader(biz_name)
            st.metric("Total Reviews", f"{m['total_reviews']:,}")
            st.metric("Weighted Avg Rating", f"{m['weighted_avg_rating']:.1f}" if m["weighted_avg_rating"] else "?")
            st.metric("Unique Products", f"{m['n_products']} ({m['n_variations']} variants)")
            st.metric("Main Product Reviews", f"{m['main_reviews']:,}")
            st.metric("Main Product Rating", f"{m['main_rating']:.1f}" if m["main_rating"] else "?")
            st.metric("Review Growth (reviews/mo)", f"{m['main_growth_per_mo']:.1f}")

    st.divider()

    # Comparison table
    w = biz_data["Wallaroo Wallets"]["metrics"]
    t = biz_data["TeacherFav"]["metrics"]

    table_data = {
        "Metric": [
            "Unique Products (Variants)",
            "Total Reviews (All Products)",
            "Weighted Avg Rating",
            "Main Product",
            "Main Product Rating",
            "Main Product Reviews",
            "Rating Trend (Main)",
            "Review Growth (Main, reviews/mo)",
            "Review Growth Direction",
            "Data Range (Main)",
            "Data Source",
        ],
        "Wallaroo Wallets": [
            f"{w['n_products']} ({w['n_variations']})",
            f"{w['total_reviews']:,}",
            f"{w['weighted_avg_rating']:.1f}" if w["weighted_avg_rating"] else "?",
            short_title(w["main_title"], 50),
            f"{w['main_rating']:.1f}" if w["main_rating"] else "?",
            f"{w['main_reviews']:,}",
            f"{w['rating_periods']['1yr']['from_rating']:.1f} → {w['rating_periods']['1yr']['to_rating']:.1f} ({w['rating_periods']['1yr']['delta']:+.2f}, {w['rating_periods']['1yr']['from_month']} → now)" if '1yr' in w.get('rating_periods', {}) else f"{w['main_r_dir']} ({w['main_early_r']:.1f} -> {w['main_recent_r']:.1f})",
            f"{w['main_growth_per_mo']:.1f} reviews/mo",
            f"{w['review_periods']['1yr']['per_month']:.1f}/mo (1yr) vs {w['review_periods']['2yr']['per_month']:.1f}/mo (2yr) — {'slowing' if w['review_periods']['1yr']['per_month'] < w['review_periods']['2yr']['per_month'] * 0.9 else 'accelerating' if w['review_periods']['1yr']['per_month'] > w['review_periods']['2yr']['per_month'] * 1.1 else 'steady'}" if '1yr' in w.get('review_periods', {}) and '2yr' in w.get('review_periods', {}) else f"{w['growth_dir']} (early {w['early_growth_mo']:.1f} -> recent {w['recent_growth_mo']:.1f} reviews/mo)",
            f"{w['main_first_date']} to {w['main_last_date']}",
            "Keepa API (complete)",
        ],
        "TeacherFav": [
            f"{t['n_products']} ({t['n_variations']})",
            f"{t['total_reviews']:,}",
            f"{t['weighted_avg_rating']:.1f}" if t["weighted_avg_rating"] else "?",
            short_title(t["main_title"], 50),
            f"{t['main_rating']:.1f}" if t["main_rating"] else "?",
            f"{t['main_reviews']:,}",
            f"{t['rating_periods']['1yr']['from_rating']:.1f} → {t['rating_periods']['1yr']['to_rating']:.1f} ({t['rating_periods']['1yr']['delta']:+.2f}, {t['rating_periods']['1yr']['from_month']} → now)" if '1yr' in t.get('rating_periods', {}) else f"{t['main_r_dir']} ({t['main_early_r']:.1f} -> {t['main_recent_r']:.1f})",
            f"{t['main_growth_per_mo']:.1f} reviews/mo",
            f"{t['review_periods']['1yr']['per_month']:.1f}/mo (1yr) vs {t['review_periods']['2yr']['per_month']:.1f}/mo (2yr) — {'slowing' if t['review_periods']['1yr']['per_month'] < t['review_periods']['2yr']['per_month'] * 0.9 else 'accelerating' if t['review_periods']['1yr']['per_month'] > t['review_periods']['2yr']['per_month'] * 1.1 else 'steady'}" if '1yr' in t.get('review_periods', {}) and '2yr' in t.get('review_periods', {}) else f"{t['growth_dir']} (early {t['early_growth_mo']:.1f} -> recent {t['recent_growth_mo']:.1f} reviews/mo)",
            f"{t['main_first_date']} to {t['main_last_date']}",
            "Keepa API (complete)",
        ],
    }
    st.dataframe(pd.DataFrame(table_data).set_index("Metric"), use_container_width=True)

    # Recent Trends comparison table
    st.divider()
    st.markdown("#### Recent Trends (Main Product)")
    st.caption("Rating, review growth, and sales rank for the main product over 6 months, 1 year, and 2 years. "
               "Sales rank: lower number = more sales; negative % = rank improved.")
    today = date.today()
    period_months = {"6mo": 6, "1yr": 12, "2yr": 24}
    for biz_name in ["Wallaroo Wallets", "TeacherFav"]:
        st.markdown(f"**{biz_name}**")
        bm = biz_data[biz_name]["metrics"]
        trend_rows = []
        for label in ["6mo", "1yr", "2yr"]:
            start = today - relativedelta(months=period_months[label])
            row = {"Period": f"Last {label} ({start.strftime('%Y-%m')} → {today.strftime('%Y-%m')})"}
            # Rating
            if label in bm.get("rating_periods", {}):
                rp = bm["rating_periods"][label]
                row["Rating"] = f"{rp['from_rating']:.1f} → {rp['to_rating']:.1f} ({rp['delta']:+.2f})"
            else:
                row["Rating"] = "—"
            # Reviews added + per month
            if label in bm.get("review_periods", {}):
                rvp = bm["review_periods"][label]
                row["Reviews"] = f"+{rvp['added']:,} ({rvp['per_month']:.1f}/mo)"
            else:
                row["Reviews"] = "—"
            # Sales Rank
            if label in bm.get("main_rank_periods", {}):
                srp = bm["main_rank_periods"][label]
                row["Sales Rank"] = f"#{srp['from_rank']:,.0f} → #{srp['to_rank']:,.0f} ({srp['pct_change']:+.0f}%)"
            else:
                row["Sales Rank"] = "—"
            trend_rows.append(row)
        if trend_rows:
            st.dataframe(pd.DataFrame(trend_rows), use_container_width=True, hide_index=True)

    # Trend charts — highlighted line charts for each time window
    st.markdown("#### Trend Charts (Main Product)")
    st.caption("Full time-series with the recent window highlighted. "
               "Annotations compare the recent window against the equivalent prior period.")

    window_options_raw = {"6mo": 6, "1yr": 12, "2yr": 24}
    window_display = []
    for wlabel, wmonths in window_options_raw.items():
        wstart = (today - relativedelta(months=wmonths)).strftime("%Y-%m")
        wend = today.strftime("%Y-%m")
        window_display.append(f"{wlabel} ({wstart} → {wend})")
    window_sel = st.selectbox("Time Window", window_display, key="trend_window_sel")
    window_label = window_sel.split(" (")[0]  # extract "6mo", "1yr", "2yr"
    months = window_options_raw[window_label]
    label = window_label

    for biz_name in ["Wallaroo Wallets", "TeacherFav"]:
        st.markdown("---")
        prods = biz_data[biz_name]["products"]
        main = max(prods, key=lambda p: p.get("reviewCount", 0)) if prods else None
        if not main:
            st.info(f"No product data available for {biz_name}.")
            continue
        rh = main.get("rating_history", [])
        ch = main.get("review_count_history", [])
        if not rh and not ch:
            st.info("No rating or review history available for this product.")
            continue
        rh_df = pd.DataFrame(rh) if rh else pd.DataFrame(columns=["date", "rating"])
        ch_df = pd.DataFrame(ch) if ch else pd.DataFrame(columns=["date", "count"])
        if not rh_df.empty:
            rh_df["date"] = pd.to_datetime(rh_df["date"])
            rh_df = rh_df.sort_values("date")
        if not ch_df.empty:
            ch_df["date"] = pd.to_datetime(ch_df["date"])
            ch_df = ch_df.sort_values("date")
        cutoff = pd.Timestamp(today - relativedelta(months=months))
        prior_start = pd.Timestamp(today - relativedelta(months=months * 2))
        cutoff_str = cutoff.strftime("%Y-%m")
        today_str = pd.Timestamp(today).strftime("%Y-%m")
        st.markdown(f"**{biz_name}** — Last {label} ({cutoff_str} → {today_str})")

        # Compute comparison stats
        # Rating: mean in recent vs prior equivalent window
        rating_annotation = ""
        if not rh_df.empty:
            recent_r = rh_df[(rh_df["date"] >= cutoff)]["rating"]
            prior_r = rh_df[(rh_df["date"] >= prior_start) & (rh_df["date"] < cutoff)]["rating"]
            r_recent_avg = round(recent_r.mean(), 2) if len(recent_r) > 0 else None
            r_prior_avg = round(prior_r.mean(), 2) if len(prior_r) > 0 else None
            parts = []
            if r_recent_avg is not None:
                parts.append(f"Recent avg: {r_recent_avg}")
            if r_prior_avg is not None:
                parts.append(f"Prior {label} avg: {r_prior_avg}")
            rating_annotation = " | ".join(parts)

        # Reviews: total added and per-month using same algorithm as metrics
        bm = biz_data[biz_name]["metrics"]
        added_recent = None
        added_prior = None
        recent_per_mo = None
        prior_per_mo = None
        ch_raw = main.get("review_count_history", [])
        if label in bm.get("review_periods", {}):
            rvp = bm["review_periods"][label]
            added_recent = rvp["added"]
            recent_per_mo = round(rvp["per_month"], 1)
        # Prior: replicate same algorithm anchored at the cutoff snapshot
        if ch_raw and len(ch_raw) >= 2:
            cutoff_dt_m = datetime.fromisoformat(ch_raw[-1]["date"]) - timedelta(days=months * 30.44)
            past_at_cutoff = [c for c in ch_raw if c["date"] <= cutoff_dt_m.isoformat()]
            if past_at_cutoff:
                anchor = past_at_cutoff[-1]
                prior_cutoff_dt = datetime.fromisoformat(anchor["date"]) - timedelta(days=months * 30.44)
                prior_past = [c for c in ch_raw if c["date"] <= prior_cutoff_dt.isoformat()]
                if prior_past:
                    ref = prior_past[-1]
                    added_prior = anchor["count"] - ref["count"]
                    span_prior = max(1, (datetime.fromisoformat(anchor["date"]) - datetime.fromisoformat(ref["date"])).days / 30.44)
                    prior_per_mo = round(added_prior / span_prior, 1)
        review_annotation = ""
        parts = []
        if added_recent is not None:
            parts.append(f"Recent: +{added_recent:,} added")
        if added_prior is not None:
            parts.append(f"Prior {label}: +{added_prior:,} added")
        review_annotation = " | ".join(parts)

        col_rating, col_reviews = st.columns(2)
        with col_rating:
            if not rh_df.empty and len(rh_df) > 0:
                fig_r = go.Figure()
                fig_r.add_trace(go.Scatter(
                    x=rh_df["date"], y=rh_df["rating"],
                    mode="lines+markers", marker=dict(size=4),
                    line=dict(color="#636EFA"), name="Rating",
                ))
                fig_r.add_vrect(
                    x0=cutoff, x1=pd.Timestamp(today),
                    fillcolor="rgba(239,85,59,0.15)", line_width=0,
                )
                fig_r.update_layout(
                    title=f"Rating", height=220,
                    margin=dict(t=50, b=30, l=50, r=20),
                    yaxis=dict(range=[3.5, 5.0]),
                    showlegend=False,
                )
                if rating_annotation:
                    fig_r.add_annotation(
                        text=rating_annotation, xref="paper", yref="paper",
                        x=0.5, y=1.08, showarrow=False,
                        font=dict(size=11, color="#555"),
                    )
                st.plotly_chart(fig_r, use_container_width=True, key=f"rating_line_{biz_name}_{label}")
            else:
                st.info("No rating history available.")
        with col_reviews:
            if not ch_df.empty and len(ch_df) > 0:
                fig_rv = go.Figure()
                fig_rv.add_trace(go.Scatter(
                    x=ch_df["date"], y=ch_df["count"],
                    mode="lines+markers", marker=dict(size=4),
                    line=dict(color="#636EFA"), name="Review Count",
                ))
                fig_rv.add_vrect(
                    x0=cutoff, x1=pd.Timestamp(today),
                    fillcolor="rgba(239,85,59,0.15)", line_width=0,
                )
                fig_rv.update_layout(
                    title=f"Review Count", height=220,
                    margin=dict(t=50, b=30, l=50, r=20),
                    showlegend=False,
                )
                if review_annotation:
                    fig_rv.add_annotation(
                        text=review_annotation, xref="paper", yref="paper",
                        x=0.5, y=1.08, showarrow=False,
                        font=dict(size=11, color="#555"),
                    )
                st.plotly_chart(fig_rv, use_container_width=True, key=f"review_line_{biz_name}_{label}")
            else:
                st.info("No review count history available.")

        # --- Comparison bar charts + frequency line chart ---
        # Compute review frequency (reviews added per interval) from cumulative
        freq_df = pd.DataFrame(columns=["date", "freq"])
        if not ch_df.empty and len(ch_df) >= 2:
            freq_df = ch_df.copy()
            freq_df["freq"] = freq_df["count"].diff()
            freq_df["days"] = freq_df["date"].diff().dt.days
            freq_df = freq_df.dropna(subset=["freq"])
            freq_df = freq_df[freq_df["days"] > 0]
            freq_df["freq"] = (freq_df["freq"] / freq_df["days"] * 30.44).round(1)
            freq_df = freq_df[freq_df["freq"] >= 0]  # drop Keepa data-correction artifacts
            freq_df = freq_df.drop(columns=["days"])

        # Rating: highlighted avg vs rest avg
        r_highlighted_avg = None
        r_rest_avg = None
        if not rh_df.empty:
            highlighted_r = rh_df[rh_df["date"] >= cutoff]["rating"]
            rest_r = rh_df[rh_df["date"] < cutoff]["rating"]
            r_highlighted_avg = round(highlighted_r.mean(), 2) if len(highlighted_r) > 0 else None
            r_rest_avg = round(rest_r.mean(), 2) if len(rest_r) > 0 else None

        # Frequency: reuse values computed above (same as table)
        f_highlighted_avg = recent_per_mo
        f_rest_avg = prior_per_mo

        prior_start_str = prior_start.strftime("%Y-%m")
        cb1, cb2, cb3 = st.columns(3)
        with cb1:
            if r_highlighted_avg is not None or r_rest_avg is not None:
                fig_rb = go.Figure()
                fig_rb.add_trace(go.Bar(
                    x=[f"Recent {label}<br>{cutoff_str}→{today_str}", f"Prior {label}<br>{prior_start_str}→{cutoff_str}"],
                    y=[r_highlighted_avg, r_rest_avg],
                    marker_color=["#EF553B", "#636EFA"],
                    text=[f"{v}" if v is not None else "" for v in [r_highlighted_avg, r_rest_avg]],
                    textposition="auto",
                ))
                fig_rb.update_layout(
                    title=f"Avg Rating (mean over {label})", height=220,
                    margin=dict(t=50, b=30, l=50, r=20),
                    yaxis=dict(range=[3.5, 5.0]),
                    showlegend=False,
                )
                st.plotly_chart(fig_rb, use_container_width=True, key=f"rating_bar_{biz_name}_{label}")
            else:
                st.info("No rating data for comparison.")
        with cb2:
            if not freq_df.empty and len(freq_df) > 0:
                fig_freq = go.Figure()
                fig_freq.add_trace(go.Scatter(
                    x=freq_df["date"], y=freq_df["freq"],
                    mode="lines+markers", marker=dict(size=4),
                    line=dict(color="#636EFA"), name="Reviews Added",
                ))
                fig_freq.add_vrect(
                    x0=cutoff, x1=pd.Timestamp(today),
                    fillcolor="rgba(239,85,59,0.15)", line_width=0,
                )
                fig_freq.update_layout(
                    title="Review Frequency (reviews/mo)", height=220,
                    margin=dict(t=50, b=30, l=50, r=20),
                    showlegend=False,
                )
                st.plotly_chart(fig_freq, use_container_width=True, key=f"freq_line_{biz_name}_{label}")
            else:
                st.info("No review frequency data available.")
        with cb3:
            # Box plot of review frequency: recent vs prior period
            recent_freq = freq_df[freq_df["date"] >= cutoff]["freq"] if not freq_df.empty else pd.Series(dtype=float)
            prior_freq = freq_df[(freq_df["date"] >= prior_start) & (freq_df["date"] < cutoff)]["freq"] if not freq_df.empty else pd.Series(dtype=float)
            if len(recent_freq) > 1 or len(prior_freq) > 1:
                fig_box = go.Figure()
                if len(prior_freq) > 1:
                    fig_box.add_trace(go.Box(
                        y=prior_freq, name=f"Prior {label}",
                        marker_color="#636EFA", boxmean=True,
                    ))
                if len(recent_freq) > 1:
                    fig_box.add_trace(go.Box(
                        y=recent_freq, name=f"Recent {label}",
                        marker_color="#EF553B", boxmean=True,
                    ))
                fig_box.update_layout(
                    title=f"Reviews/Mo Distribution ({label})", height=220,
                    margin=dict(t=50, b=30, l=50, r=20),
                    yaxis=dict(rangemode="tozero"),
                    showlegend=False,
                )
                st.plotly_chart(fig_box, use_container_width=True, key=f"freq_box_{biz_name}_{label}")
            else:
                st.info("Not enough data points for box plot.")

    st.caption(
        "**How chart values are calculated:** "
        "• **Avg Rating** = mean of all Keepa rating snapshots within the window. "
        "• **Review Frequency line** = reviews added between consecutive Keepa snapshots, normalized to reviews/month "
        "(diff ÷ days × 30.44). Negative diffs from Keepa data corrections are excluded. "
        "• **Reviews/Mo Distribution box plot** = spread of per-snapshot reviews/month values; diamond = mean, line = median. "
        "Shows whether review velocity is steady or spiky. "
        "• **Recent** = the selected time window ending today. **Prior** = the same-length window immediately before it."
    )

    with st.expander("Metric Definitions, Review Integrity & Date Range Impact", expanded=False):
        for biz_name in ["Wallaroo Wallets", "TeacherFav"]:
            m = biz_data[biz_name]["metrics"]
            integrity = compute_integrity_signals(biz_data[biz_name]["products"])
            flagged = [s for s in integrity if not s["looks_organic"]]

            st.markdown(f"**{biz_name}**")
            st.markdown(
                f"Data covers {m['main_first_date']} to {m['main_last_date']} (main product). "
                f"Review growth is {m['main_growth_per_mo']:.1f} reviews/mo — "
                f"{'this is a high rate of new reviews for the category.' if m['main_growth_per_mo'] > 30 else 'a moderate rate of new reviews.'}"
            )
            if not flagged:
                st.markdown(f"No manipulation signals detected across {len(integrity)} products. Review growth looks organic.")
            else:
                for s in flagged:
                    st.markdown(f"- **{s['name']}**: {'; '.join(s['red_flags'])}")
            st.markdown("")

        st.markdown("""
**Reading the comparison table:**
- **Weighted Avg Rating** — products with more reviews have more weight. A 4.6 product with 7,000 reviews dominates over a 3.4 product with 16 reviews.
- **Rating Trend** — STABLE/IMPROVING/DECLINING compares first half vs second half of the data. The numbers in parentheses show the actual average shift.
- **Review Growth Direction** — ACCELERATING means the product is gaining reviews faster recently; DECELERATING means it's slowing down.
- **Data Range** — the full period covered by Keepa data for the main product.
""")


def page_business_analysis(biz_data, biz_name):
    products = biz_data[biz_name]["products"]
    metrics = biz_data[biz_name]["metrics"]
    main = products[0]

    st.header(f"{biz_name} — Review Analysis")
    st.caption("Based on Keepa API data | Full historical range")

    # KPI row
    cols = st.columns(6)
    cols[0].metric("Products", f"{metrics['n_products']} ({metrics['n_variations']} var)")
    cols[1].metric("Total Reviews", f"{metrics['total_reviews']:,}")
    cols[2].metric("Avg Rating", f"{metrics['weighted_avg_rating']:.1f}" if metrics["weighted_avg_rating"] else "?")
    cols[3].metric("Review Growth (reviews/mo)", f"{metrics['main_growth_per_mo']:.1f}")
    cols[4].metric("Rating Trend", metrics["main_r_dir"])
    if metrics["main_rank_recent"] is not None:
        current_rank = metrics["main_rank_recent"]

        # Show 1yr change as delta if available, else 6mo
        rank_periods = metrics["main_rank_periods"]
        rank_delta_text = None
        compare_period = None
        for period_key in ["1yr", "6mo", "2yr"]:
            if period_key in rank_periods:
                p = rank_periods[period_key]
                pct = p["pct_change"]
                compare_period = period_key
                rank_delta_text = f"{pct:+.0f}% vs {period_key} ago (was #{p['from_rank']:,.0f})"
                rank_delta_color = "inverse"  # inverse: negative shown green, positive shown red
                break
        if rank_delta_text:
            cols[5].metric("Main Product Sales Rank", f"#{current_rank:,.0f}", delta=rank_delta_text, delta_color=rank_delta_color)
        else:
            cols[5].metric("Main Product Sales Rank", f"#{current_rank:,.0f}")
        # Factual explanation: what the number means and direction
        explanation = f"#{current_rank:,.0f} means {current_rank - 1:,.0f} products in this category sell more."
        if rank_delta_text and compare_period:
            p = rank_periods[compare_period]
            from_rank = p["from_rank"]
            if p["pct_change"] > 20:
                explanation += f" Was #{from_rank:,.0f} {compare_period} ago — rank number went up, meaning fewer sales now."
            elif p["pct_change"] < -20:
                explanation += f" Was #{from_rank:,.0f} {compare_period} ago — rank number went down, meaning more sales now."
            else:
                explanation += f" Was #{from_rank:,.0f} {compare_period} ago — roughly the same sales level."
        cols[5].caption(explanation)
    else:
        cols[5].metric("Main Product Sales Rank", "No data")

    st.divider()

    # Review integrity + context
    integrity = compute_integrity_signals(products)
    main_sig = integrity[0]  # main product (sorted by reviews desc)

    with st.expander("Metric Definitions, Review Integrity & Date Range Impact", expanded=False):
        st.markdown(f"""
**You're looking at** a summary of {biz_name}'s Amazon review data from Keepa, using the full historical data range.

**Reading the numbers:**
- **Products** — number of unique products (deduplicated by parent ASIN). The number in parentheses is total color/style variants sharing those review pools.
- **Total Reviews** — sum of review counts across all products at the latest data point.
- **Avg Rating** — weighted average rating across all products, weighted by review count (so a product with 7,000 reviews matters more than one with 100).
- **Review Growth** — how many new reviews the main product gains per month on average. Higher = more customers buying.
- **Rating Trend** — compares the average rating in the first half vs second half of the data. STABLE means quality hasn't changed; DECLINING means recent ratings are lower.
- **Main Product Sales Rank** — where the main product ranks in sales compared to every other product in the same Amazon category. Lower number = more sales. For example, rank 3,000 means only 2,999 products sell more in that category.

**Review Integrity Assessment**
""")

        for sig in integrity:
            if sig["looks_organic"] and not sig["red_flags"]:
                st.markdown(f"- **{sig['name']}**: Reviews look organic. Steady growth, no Amazon purges, stable ratings{' (std dev: ' + str(sig['rating_std']) + ')' if sig['rating_std'] is not None else ''}.")
            elif sig["looks_organic"]:
                st.markdown(f"- **{sig['name']}**: Mostly organic. {'; '.join(sig['red_flags'])}.")
            else:
                st.markdown(f"- **{sig['name']}**: **Review flags detected.** {'; '.join(sig['red_flags'])}.")

        flagged = [s for s in integrity if not s["looks_organic"]]
        if not flagged:
            st.success("Overall: No significant manipulation signals detected across the portfolio. Review growth patterns are consistent and Amazon has not performed major review purges.")
        else:
            st.warning(f"Overall: {len(flagged)} product(s) have review flags worth investigating. See details above.")

        st.markdown(f"""
**How the date range affects these values:**
Changing the date range in the sidebar recalculates everything on this page. Narrowing to the last 1-2 years isolates recent performance — useful for spotting if a once-strong product is now declining. The full range gives the historical average, which smooths out short-term noise but can hide recent trends.
""")

    tab1, tab2, tab3, tab4 = st.tabs(["Main Product", "Portfolio", "All Products", "Product Table"])

    # ─── 1. Main Product Only ───
    with tab1:
        st.subheader(f"Main Product: {short_title(main['title'])}")
        st.caption(
            f"ASIN: [{main['representative_asin']}]({amazon_link(main['representative_asin'])}) | "
            f"[Keepa]({keepa_link(main['representative_asin'])}) | "
            f"{main['shared_reviews']:,} reviews | {main['num_variations']} variants"
        )

        # Data range for this product
        main_ch = main["review_count_history"]
        main_rh = main["rating_history"]
        if main_ch:
            st.caption(
                f"Data range: {main_ch[0]['date'][:7]} to {main_ch[-1]['date'][:7]} "
                f"({len(main_ch)} review observations, {len(main_rh)} rating observations)"
            )
        if main_sig["red_flags"]:
            st.caption(f"Review flags: {'; '.join(main_sig['red_flags'])}")
        else:
            st.caption("Review integrity: No manipulation signals detected — steady growth, no Amazon purges.")

        fig = make_rating_chart(
            [main],
            f"Rating History — {product_name(main, short=True)}",
            colors=["#00d4aa"],
        )
        st.plotly_chart(fig, use_container_width=True)
        analysis = _analyze_rating([main])
        if analysis:
            _ai_caption(analysis)

        fig = make_review_count_chart(
            [main],
            f"Review Count Growth — {product_name(main, short=True)}",
            colors=["#00d4aa"],
        )
        st.plotly_chart(fig, use_container_width=True)
        analysis = _analyze_review_count([main])
        if analysis:
            _ai_caption(analysis)

        # Price history chart
        if main.get("marketplace_price_history"):
            fig = make_price_chart(
                [main],
                f"Price History — {product_name(main, short=True)}",
                colors=["#00d4aa"],
            )
            st.plotly_chart(fig, use_container_width=True)
            analysis = _analyze_price(main)
            if analysis:
                _ai_caption(analysis)

            # Price correlation analysis
            col_pv, col_pr = st.columns(2)
            with col_pv:
                fig_pv = _make_price_vs_velocity_chart(
                    main, f"Price vs Review Velocity")
                if fig_pv:
                    st.plotly_chart(fig_pv, use_container_width=True)
            with col_pr:
                fig_pr = _make_price_vs_rating_chart(
                    main, f"Price vs Rating")
                if fig_pr:
                    st.plotly_chart(fig_pr, use_container_width=True)

            # Scatter plots with regression
            fig_sc_rating, fig_sc_velocity = _make_price_scatter(main)
            if fig_sc_rating or fig_sc_velocity:
                col_sr, col_sv = st.columns(2)
                if fig_sc_rating:
                    with col_sr:
                        st.plotly_chart(fig_sc_rating, use_container_width=True)
                if fig_sc_velocity:
                    with col_sv:
                        st.plotly_chart(fig_sc_velocity, use_container_width=True)

            price_insight = _compute_price_insight(main)
            if price_insight:
                st.info(price_insight)

        # Reviews per month velocity chart
        if len(main_ch) >= 4:
            monthly_data = {}
            for c in main_ch:
                ym = c["date"][:7]
                monthly_data[ym] = c["count"]
            months_sorted = sorted(monthly_data.keys())
            velocity_months = []
            velocity_values = []
            for i in range(1, len(months_sorted)):
                prev_m, curr_m = months_sorted[i - 1], months_sorted[i]
                prev_dt = datetime.strptime(prev_m, "%Y-%m")
                curr_dt = datetime.strptime(curr_m, "%Y-%m")
                days_between = max(1, (curr_dt - prev_dt).days)
                new_reviews = monthly_data[curr_m] - monthly_data[prev_m]
                per_month = new_reviews / (days_between / 30.44)
                velocity_months.append(curr_dt)
                velocity_values.append(round(per_month, 1))

            fig_vel = go.Figure()
            fig_vel.add_trace(go.Bar(
                x=velocity_months,
                y=velocity_values,
                marker_color=["#00d4aa" if v >= 0 else "#ff6b6b" for v in velocity_values],
            ))
            fig_vel.update_layout(
                title=dict(text=f"New Reviews per Month — {product_name(main, short=True)}<br><sup>Rate of new reviews added each month (not cumulative). Negative bars = Amazon removed reviews that month.</sup>"),
                yaxis_title="New Reviews / Month",
                height=400,
                margin=dict(l=60, r=20, t=70, b=40),
            )
            st.plotly_chart(fig_vel, use_container_width=True)
            analysis = _analyze_velocity(velocity_values)
            if analysis:
                _ai_caption(analysis)

        # Sales rank chart for main product
        rank_data = main.get("monthly_avg_rank", {})
        if rank_data and len(rank_data) >= 2:
            fig_rank = make_sales_rank_chart([main], f"Sales Rank — {product_name(main, short=True)}")
            st.plotly_chart(fig_rank, use_container_width=True)
            fig_rank_zoom = make_sales_rank_chart([main], f"Sales Rank (Zoomed In) — {product_name(main, short=True)}", zoomed=True)
            st.plotly_chart(fig_rank_zoom, use_container_width=True)
            analysis = _analyze_sales_rank([main])
            if analysis:
                _ai_caption(analysis)

        render_notable_events([main], product_name(main, short=True))

    # ─── 2. All Products Combined ───
    with tab2:
        st.subheader("All Products Combined (Portfolio Performance)")

        fig = make_combined_chart(
            products, "rating",
            f"{biz_name} — Weighted Avg Rating (All Products Combined)",
        )
        st.plotly_chart(fig, use_container_width=True)
        analysis = _analyze_combined_rating(products)
        if analysis:
            _ai_caption(analysis)

        fig = make_combined_chart(
            products, "reviews",
            f"{biz_name} — Total Reviews (All Products Combined)",
        )
        st.plotly_chart(fig, use_container_width=True)
        analysis = _analyze_combined_reviews(products)
        if analysis:
            _ai_caption(analysis)

        render_notable_events(products, "Portfolio")

    # ─── 3. All Individual Products ───
    with tab3:
        st.subheader("All Individual Products (Comprehensive Breakdown)")

        fig = make_rating_chart(products, f"{biz_name} — Rating History (All Products)")
        st.plotly_chart(fig, use_container_width=True)
        analysis = _analyze_rating(products)
        if analysis:
            _ai_caption(analysis)

        fig = make_review_count_chart(products, f"{biz_name} — Review Count Growth (All Products)")
        st.plotly_chart(fig, use_container_width=True)
        analysis = _analyze_review_count(products)
        if analysis:
            _ai_caption(analysis)

        # Sales rank
        fig = make_sales_rank_chart(products, f"{biz_name} — Monthly Avg Sales Rank")
        st.plotly_chart(fig, use_container_width=True)
        fig_zoom = make_sales_rank_chart(products, f"{biz_name} — Monthly Avg Sales Rank (Zoomed In)", zoomed=True)
        st.plotly_chart(fig_zoom, use_container_width=True)
        analysis = _analyze_sales_rank(products)
        if analysis:
            _ai_caption(analysis)

        # Price history
        products_with_price = [p for p in products if p.get("marketplace_price_history")]
        if products_with_price:
            fig = make_price_chart(products_with_price, f"{biz_name} — Price History (All Products)")
            st.plotly_chart(fig, use_container_width=True)
            # AI analysis for portfolio price
            price_ranges = []
            for pp in products_with_price:
                pph = pp.get("marketplace_price_history", [])
                if pph:
                    price_ranges.append((product_name(pp, short=True), pph[-1]["price"]))
            if price_ranges:
                prices_sorted = sorted(price_ranges, key=lambda x: -x[1])
                highest = prices_sorted[0]
                lowest = prices_sorted[-1]
                st.caption(
                    f"*AI analysis (Claude):* Across the {biz_name} portfolio, "
                    f"prices range from \\${lowest[1]:.2f} ({lowest[0]}) to \\${highest[1]:.2f} ({highest[0]}). "
                    f"{'Wide price spread means the products serve different customer segments — each price level faces different competitors.' if highest[1] / max(lowest[1], 0.01) > 2 else 'Products are priced in a similar range, meaning they compete in the same part of the market.'}"
                )

        render_notable_events(products, "All Products")

    # ─── Product Breakdown Table ───
    with tab4:
        st.subheader("Product Breakdown Table")

        rows = []
        for p in products:
            rh = p["rating_history"]
            ch = p["review_count_history"]
            rating = f"{rh[-1]['rating']:.1f}" if rh else "?"

            if len(ch) >= 2:
                first_date = datetime.fromisoformat(ch[0]["date"])
                last_date = datetime.fromisoformat(ch[-1]["date"])
                months = max(1, (last_date - first_date).days / 30.44)
                growth = (ch[-1]["count"] - ch[0]["count"]) / months
                growth_str = f"{growth:.1f}"
                since = ch[0]["date"][:7]
            else:
                growth_str = "?"
                since = "?"

            if len(rh) >= 4:
                mid = len(rh) // 2
                early = sum(r["rating"] for r in rh[:mid]) / mid
                recent = sum(r["rating"] for r in rh[mid:]) / (len(rh) - mid)
                trend = "STABLE" if abs(recent - early) <= 0.05 else ("UP" if recent > early else "DOWN")
            else:
                trend = "?"

            # Latest marketplace price
            mph = p.get("marketplace_price_history", [])
            price_str = f"${mph[-1]['price']:.2f}" if mph else "?"

            rows.append({
                "ASIN": p["representative_asin"],
                "Product": product_name(p, short=True),
                "Amazon": amazon_link(p["representative_asin"]),
                "Keepa": keepa_link(p["representative_asin"]),
                "Rating": rating,
                "Price": price_str,
                "Reviews": f"{p['shared_reviews']:,}",
                "Variants": p["num_variations"],
                "Reviews/mo": growth_str,
                "Rating Trend": trend,
                "Data Range": f"{since} to {ch[-1]['date'][:7]}" if len(ch) >= 2 else since,
            })

        df = pd.DataFrame(rows)
        st.dataframe(
            df,
            column_config={
                "Amazon": st.column_config.LinkColumn("Amazon", display_text="View"),
                "Keepa": st.column_config.LinkColumn("Keepa", display_text="View"),
            },
            use_container_width=True,
            hide_index=True,
        )

        with st.expander("ℹ️ Column Definitions"):
            st.markdown("""
| Column | Definition |
|--------|-----------|
| **ASIN** | Amazon Standard Identification Number — unique product identifier |
| **Product** | Short display name for the product |
| **Amazon / Keepa** | Links to the product's Amazon listing and Keepa tracking page |
| **Rating** | Most recent star rating from Keepa data (last data point in rating history). Keepa stores ratings as integers (e.g. 43 = 4.3 stars), divided by 10 for display. Shows "?" if no rating data exists. |
| **Reviews** | Total review count shared across all variants of this product (from `shared_reviews`). This is the review count at the most recent Keepa data point. |
| **Variants** | Number of color/style variations sharing the same review pool (from `num_variations`) |
| **Reviews/mo** | Average new reviews per month = `(last_count - first_count) / months_between`. Calculated over the full data range. Shows "?" if fewer than 2 data points. |
| **Rating Trend** | Compares average rating in the first half vs second half of rating history. **UP** = recent avg > early avg by >0.05, **DOWN** = recent avg < early avg by >0.05, **STABLE** = within ±0.05. Shows "?" if fewer than 4 rating data points. |
| **Data Range** | Date range of Keepa review count data (first to last observation), formatted as YYYY-MM. |

---

**Example:** A product with 10 rating samples \\[4.5, 4.5, 4.4, 4.4, 4.3, 4.2, 4.2, 4.1, 4.1, 4.0\\]:
- **Rating** = 4.0 (last data point)
- **Rating Trend** = early half avg (4.5+4.5+4.4+4.4+4.3)/5 = 4.42, recent half avg (4.2+4.2+4.1+4.1+4.0)/5 = 4.12 → difference is −0.30 → **DOWN**
- If first review count was 100 on 2023-01 and last was 220 on 2024-01 (12 months): **Reviews/mo** = (220−100)/12 = 10.0

---

**How values are computed:**
All values use the full historical data range. Rating uses the last data point. Reviews/mo divides review growth across the total months of data. Rating Trend splits the data in half. If a product has no data, values show "?".
""")


def page_evaluation(biz_data, biz_name):
    products = biz_data[biz_name]["products"]
    m = biz_data[biz_name]["metrics"]

    st.header(f"{biz_name} — Acquisition Evaluation")
    st.caption("Based on Keepa API data | Full historical range. Verify financials independently.")

    main = products[0]
    main_ch = main["review_count_history"]
    if main_ch:
        st.caption(
            f"Showing data from {main_ch[0]['date'][:7]} to {main_ch[-1]['date'][:7]} "
            f"(main product: {product_name(main, short=True)}). "
            f"Adjust the date range in the sidebar to change the evaluation window."
        )

    if biz_name == "Wallaroo Wallets":
        st.markdown(
            "Listing: [Flippa](https://flippa.com/12260386-amazon-s-favorite-leather-phone-wallet-since-2016-"
            "hundreds-of-thousands-of-customers-over-7-000-reviews-for-the-hero-sku-with-an-average-4-6-star-rating)"
        )
    else:
        st.markdown(
            "Listing: [Empire Flippers #92221](https://app.empireflippers.com/listing/unlocked/92221)"
        )

    # Helper: build rank bullet points with specific time windows
    def _rank_bullets(metrics):
        """Return (pros_list, cons_list) for sales rank with time-specific data."""
        pros, cons = [], []
        periods = metrics.get("main_rank_periods", {})
        current = metrics.get("main_rank_recent")
        if current is None or not periods:
            return pros, cons
        for label in ["6mo", "1yr", "2yr"]:
            if label not in periods:
                continue
            p = periods[label]
            pct = p["pct_change"]
            if abs(pct) < 20:
                direction = "stable"
            elif pct < 0:
                direction = "improved"
            else:
                direction = "worsened"
            text = (
                f"Main product sales rank {direction}: ~{p['from_rank']:,.0f} → ~{p['to_rank']:,.0f} "
                f"({pct:+.0f}%) over the last {label} ({p['from_month']} to {metrics['main_rank_recent_month']})"
            )
            if direction == "improved":
                pros.append(text)
            elif direction == "worsened":
                cons.append(text)
            # stable: not notable enough for pro/con
        return pros, cons

    def _rank_summary_note(metrics):
        """Return a factual rank summary sentence for the AI assessment."""
        periods = metrics.get("main_rank_periods", {})
        current = metrics.get("main_rank_recent")
        if current is None:
            return ""
        parts = []
        for label in ["1yr", "6mo"]:
            if label in periods:
                p = periods[label]
                pct = p["pct_change"]
                parts.append(f"{pct:+.0f}% vs {label} ago (was ~{p['from_rank']:,.0f})")
        rank_context = "; ".join(parts) if parts else "no prior data to compare"
        return (
            f" Sales rank is currently #{current:,.0f} ({rank_context}). "
            f"This means {current - 1:,.0f} products in the category sell more."
        )

    def _rating_trend_bullets(metrics):
        """Return (pros, cons) for rating changes over time windows."""
        pros, cons = [], []
        for label in ["1yr", "6mo"]:
            if label not in metrics.get("rating_periods", {}):
                continue
            rp = metrics["rating_periods"][label]
            delta = rp["delta"]
            if abs(delta) < 0.05:
                continue  # not notable
            text = f"Rating {rp['from_rating']:.1f} → {rp['to_rating']:.1f} ({delta:+.2f}) over last {label} ({rp['from_month']} to now)"
            if delta > 0.1:
                pros.append(text)
            elif delta < -0.1:
                cons.append(text)
        return pros, cons

    def _review_trend_bullets(metrics):
        """Return (pros, cons) for review growth changes over time windows."""
        pros, cons = [], []
        periods = metrics.get("review_periods", {})
        # Compare 1yr rate vs 2yr rate to detect acceleration/deceleration
        if "1yr" in periods and "2yr" in periods:
            recent_rate = periods["1yr"]["per_month"]
            longer_rate = periods["2yr"]["per_month"]
            if recent_rate < longer_rate * 0.7:
                cons.append(
                    f"Review growth slowing: {recent_rate:.0f} reviews/mo over last 1yr vs {longer_rate:.0f}/mo over last 2yr — fewer new customers"
                )
            elif recent_rate > longer_rate * 1.3:
                pros.append(
                    f"Review growth accelerating: {recent_rate:.0f} reviews/mo over last 1yr vs {longer_rate:.0f}/mo over last 2yr — more new customers"
                )
        # Show 6mo rate if notably different from 1yr
        if "6mo" in periods and "1yr" in periods:
            rate_6mo = periods["6mo"]["per_month"]
            rate_1yr = periods["1yr"]["per_month"]
            if rate_6mo < rate_1yr * 0.7:
                cons.append(
                    f"Recent slowdown: {rate_6mo:.0f} reviews/mo over last 6mo vs {rate_1yr:.0f}/mo over last 1yr"
                )
            elif rate_6mo > rate_1yr * 1.3:
                pros.append(
                    f"Recent acceleration: {rate_6mo:.0f} reviews/mo over last 6mo vs {rate_1yr:.0f}/mo over last 1yr"
                )
        return pros, cons

    if biz_name == "Wallaroo Wallets":
        verdict = "CAUTIOUSLY POSITIVE"
        verdict_emoji = "🟢"
        rank_pros, rank_cons = _rank_bullets(m)
        rating_pros, rating_cons = _rating_trend_bullets(m)
        review_pros, review_cons = _review_trend_bullets(m)
        pros = [
            f"10-year brand, 4.6 stars across {m['main_reviews']:,} reviews — exceptional longevity",
            f"Rating STABLE at 4.4-4.6 over full history — no quality degradation",
            f"Main product growing at {m['main_growth_per_mo']:.0f} new reviews/month consistently",
            f"{m['n_variations']} color variants from just {m['n_products']} products = efficient SKU strategy",
            "Ultra-low unit cost ($1.65-1.81) = massive tariff buffer",
            "Category leader in phone wallets — well-established with strong sales history",
        ] + rank_pros + rating_pros + review_pros
        cons = [
            f"Only {m['n_products']} unique products — high concentration in Phone Wallet ({m['main_reviews']:,}/{m['total_reviews']:,} reviews)",
            "MagSafe & Wristlet trending DOWN (4.5->4.1, 4.7->4.2) — newer products struggling",
            "No P&L data (Flippa listing) — cannot verify financial claims",
            "Competitive market: phone wallets have low barriers to entry",
            f"Card Holder & Wristlet have low volume ({159+114} reviews combined) — limited traction",
        ] + rank_cons + rating_cons + review_cons
        questions = [
            "What are the actual revenue and profit numbers? Need P&L or bank statements",
            "Why are newer products (MagSafe, Wristlet) rated lower? Quality or expectations issue?",
            "What is the product cost breakdown including tariffs?",
            "What % of revenue comes from the main Phone Wallet vs other products?",
            "Is there a supplier dependency risk? Backup manufacturers?",
        ]
        rank_note = _rank_summary_note(m)
        summary = (
            f"The main Phone Wallet has a rock-solid 4.6 rating across {m['main_reviews']:,} reviews with stable "
            f"trends — a 10-year track record of consistent quality. However, the business is essentially a "
            f"one-product company — the Phone Wallet accounts for {m['main_reviews']/m['total_reviews']*100:.0f}% "
            f"of all reviews. Newer products (MagSafe at 4.2, Wristlet at 4.2) show declining ratings.{rank_note} "
            f"The core product is strong but diversification is weak. CRITICAL: No P&L available — cannot make a "
            f"financial assessment. Proceed only with verified financials."
        )
    else:
        verdict = "MIXED — STRONG CORE, WEAK TAILS"
        verdict_emoji = "🟡"
        rank_pros, rank_cons = _rank_bullets(m)
        rating_pros, rating_cons = _rating_trend_bullets(m)
        review_pros, review_cons = _review_trend_bullets(m)
        # Check for products with badly worsening rank (with time context)
        rank_cons_extra = []
        for p in products[1:]:
            p_rank = p.get("monthly_avg_rank", {})
            if p_rank and len(p_rank) >= 2:
                p_months = sorted(p_rank.keys())
                p_early, p_recent = p_rank[p_months[0]], p_rank[p_months[-1]]
                if p_recent > p_early * 3 and p_recent > 100_000:
                    rank_cons_extra.append(
                        f"{product_name(p, short=True)} sales rank went from ~{p_early:,.0f} to ~{p_recent:,.0f} "
                        f"({p_months[0]} to {p_months[-1]}) — barely selling"
                    )
        pros = [
            f"Core Sand Timer line at 4.4 stars with {products[0]['shared_reviews']:,} reviews — market leader",
            f"Toothbrush Timer at 4.6 stars ({products[1]['shared_reviews']:,} reviews) — STRONGER than initially reported",
            f"16 color variants of main timer = wide selection that's hard for new competitors to match",
            "P&L verified: $46K/mo revenue, 10% margin via Empire Flippers",
            f"Review growth at {m['main_growth_per_mo']:.0f} new reviews/month on main product — steady stream of new buyers",
            "Two strong products (Sand Timer 4.4, Toothbrush Timer 4.6) — diversified within niche",
        ] + rank_pros + rating_pros + review_pros
        cons = [
            f"Rating trend slightly DOWN on main product ({m['main_early_r']:.1f} -> {m['main_recent_r']:.1f})",
            f"60-Minute Timer at 3.4 stars (16 reviews) — worst performing product",
            "10% margin is thin — vulnerable to cost increases or ad spend pressure",
            f"Review growth {m['growth_dir'].lower()} — early {m['early_growth_mo']:.0f} reviews/mo vs recent {m['recent_growth_mo']:.0f} reviews/mo",
            f"Dashboard: FAILS age (<5yr) and multiple (>30x). 2 FAIL criteria on acquisition checklist",
        ] + rank_cons + rank_cons_extra + rating_cons + review_cons
        questions = [
            "Can the 60-Minute Timer (3.4 stars) be discontinued or redesigned?",
            "Why is margin only 10% on $46K/mo revenue? Where are costs going?",
            "Can TaCOS / Total Ad Cost of Sales (12.9%) be reduced without losing rank?",
            "What is the defect/return rate for the Sand Timer line?",
            "Would seller accept $100K-120K (19-23x) given the 2 FAIL criteria?",
        ]
        rank_note = _rank_summary_note(m)
        if rank_cons_extra:
            rank_note += f" {len(rank_cons_extra)} product(s) have seen their sales rank collapse, suggesting they are barely selling."
        summary = (
            f"The Toothbrush Timer at 4.6 stars is the strongest product in the portfolio. The core Sand Timer "
            f"line at 4.4 stars with {products[0]['shared_reviews']:,} reviews is solid. The only weak spot is the "
            f"60-Minute Timer at 3.4 stars with just 16 reviews.{rank_note} Overall product quality is strong across the "
            f"main catalog. However, financial concerns remain: 10% margin is thin, and the listing fails 2 acquisition "
            f"criteria (age <5yr, multiple >30x). Worth pursuing at a lower price."
        )

    st.markdown(f"### {verdict_emoji} Verdict: **{verdict}**")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### Pros")
        for p in pros:
            st.markdown(f"- {p}")
    with col2:
        st.markdown("#### Cons")
        for c in cons:
            st.markdown(f"- {c}")

    st.markdown("#### Key Questions for Seller")
    for q in questions:
        st.markdown(f"- {q}")

    st.divider()

    st.markdown("#### AI Assessment")
    st.info(summary)

    # Trends table: rating, reviews, sales rank over 6mo / 1yr / 2yr
    st.divider()
    st.markdown("#### Main Product Trends by Time Window")
    st.caption(
        "How the main product's key metrics changed over the last 6 months, 1 year, and 2 years. "
        "Rating = star rating (out of 5). Reviews/mo = average new reviews added per month in that period. "
        "Sales rank = position in category (lower = more sales; negative % change = rank improved)."
    )
    trend_rows = []
    for label in ["6mo", "1yr", "2yr"]:
        row = {"Period": f"Last {label}"}
        # Rating
        if label in m.get("rating_periods", {}):
            rp = m["rating_periods"][label]
            row["Rating"] = f"{rp['from_rating']:.1f} → {rp['to_rating']:.1f} ({rp['delta']:+.2f})"
        else:
            row["Rating"] = "—"
        # Reviews per month
        if label in m.get("review_periods", {}):
            rvp = m["review_periods"][label]
            row["Reviews Added"] = f"{rvp['added']:,}"
            row["Reviews/mo"] = f"{rvp['per_month']:.1f}"
        else:
            row["Reviews Added"] = "—"
            row["Reviews/mo"] = "—"
        # Sales rank
        if label in m.get("main_rank_periods", {}):
            srp = m["main_rank_periods"][label]
            row["Sales Rank"] = f"#{srp['from_rank']:,.0f} → #{srp['to_rank']:,.0f} ({srp['pct_change']:+.0f}%)"
        else:
            row["Sales Rank"] = "—"
        trend_rows.append(row)
    if trend_rows:
        st.dataframe(pd.DataFrame(trend_rows), use_container_width=True, hide_index=True)

    # Wallaroo seller-reported financials (no P&L available, only summary table)
    if biz_name == "Wallaroo Wallets":
        st.divider()
        st.markdown("#### Seller-Reported Financials")
        st.warning(
            "⚠️ **Unverified data** — The seller provided this summary but has not shared a P&L, "
            "bank statements, or Amazon payout reports. Verify independently before relying on these numbers."
        )
        wallaroo_fin = pd.DataFrame({
            "": ["Gross Product Sales", "Net Revenue", "Total Amazon Fees", "FBA Reimbursements"],
            "2021 (Sep-Dec)": ["$225,514", "$214,864", "($78,328)", "$1,426"],
            "2022": ["$859,439", "$828,268", "($319,738)", "$8,780"],
            "2023": ["$738,150", "$696,466", "($288,554)", "$13,124"],
            "2024": ["$469,138", "$446,495", "($178,623)", "$5,753"],
        })
        st.dataframe(wallaroo_fin, use_container_width=True, hide_index=True)
        st.markdown(
            "**Notable trends:**\n"
            "- Revenue peaked in 2022 ($859K gross) and has declined since: "
            "-14% in 2023, -36% in 2024\n"
            "- Amazon fees consistently ~38-41% of net revenue\n"
            "- 2021 is partial year (Sep-Dec only) — annualized would be ~$676K gross\n"
            "- No COGS or profit data provided — margin is unknown"
        )

    # P&L Financial Data (if available)
    pnl_raw = load_pnl(biz_name)
    if pnl_raw:
        # Filter P&L to selected date range
        start_ym = st.session_state.get("date_start", date(2020, 1, 1))
        end_ym = st.session_state.get("date_end", date.today())
        if isinstance(start_ym, date):
            start_ym = start_ym.strftime("%Y-%m")
        if isinstance(end_ym, date):
            end_ym = end_ym.strftime("%Y-%m")
        pnl = {}
        for key, series in pnl_raw.items():
            pnl[key] = {m: v for m, v in series.items() if start_ym <= m <= end_ym}

        st.divider()
        st.markdown("#### Financial Data (Empire Flippers P&L)")

        # Revenue & Net Income chart
        rev = pnl.get("Total Revenue", {})
        net = pnl.get("Net Income", {})
        if rev:
            months_sorted = sorted(rev.keys())
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=[datetime.strptime(m, "%Y-%m") for m in months_sorted],
                y=[rev[m] for m in months_sorted],
                name="Revenue",
                marker_color="#3498db",
            ))
            if net:
                fig.add_trace(go.Bar(
                    x=[datetime.strptime(m, "%Y-%m") for m in months_sorted if m in net],
                    y=[net[m] for m in months_sorted if m in net],
                    name="Net Income",
                    marker_color="#2ecc71",
                ))
            fig.update_layout(
                title=dict(text="Monthly Revenue & Net Income<br><sup>Source: Empire Flippers P&L (Listing #92221)</sup>"),
                yaxis_title="USD",
                height=400,
                barmode="group",
                margin=dict(l=60, r=20, t=50, b=40),
            )
            st.plotly_chart(fig, use_container_width=True)
            # AI analysis for revenue chart
            rev_vals = [rev[m] for m in months_sorted]
            net_vals = [net.get(m, 0) for m in months_sorted]
            recent_rev = rev_vals[-3:] if len(rev_vals) >= 3 else rev_vals
            avg_recent_rev = sum(recent_rev) / len(recent_rev) if recent_rev else 0
            recent_net = net_vals[-3:] if len(net_vals) >= 3 else net_vals
            avg_recent_net = sum(recent_net) / len(recent_net) if recent_net else 0
            margin_pct = (avg_recent_net / avg_recent_rev * 100) if avg_recent_rev else 0
            if rev_vals[-1] < rev_vals[0] * 0.8:
                rev_decline = (1 - rev_vals[-1] / rev_vals[0]) * 100
                rev_stake = f"Revenue declined {rev_decline:.0f}% over this period — at this rate, the business generates \\${rev_vals[-1]:,.0f}/mo, which directly reduces the earnings base used to calculate acquisition multiples."
            else:
                rev_stake = f"Revenue is stable or growing over this period, supporting the current earnings multiple."
            _ai_caption(
                f"Recent 3-month average: \\${avg_recent_rev:,.0f}/mo revenue, \\${avg_recent_net:,.0f}/mo net income ({margin_pct:.0f}% margin). {rev_stake}"
            )

        # Margin chart
        margin = pnl.get("Net Profit Margin (12mo trailing)", {})
        tacos = pnl.get("TaCOS (Total Ad Cost of Sales)", {})
        if margin:
            months_sorted = sorted(margin.keys())
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=[datetime.strptime(m, "%Y-%m") for m in months_sorted],
                y=[margin[m] for m in months_sorted],
                name="Net Profit Margin (12mo trailing)",
                mode="lines+markers",
                line=dict(width=2, color="#2ecc71"),
            ))
            if tacos:
                fig.add_trace(go.Scatter(
                    x=[datetime.strptime(m, "%Y-%m") for m in months_sorted if m in tacos],
                    y=[tacos[m] for m in months_sorted if m in tacos],
                    name="TaCOS (Total Ad Cost of Sales)",
                    mode="lines+markers",
                    line=dict(width=2, color="#e74c3c"),
                ))
            fig.update_layout(
                title=dict(text="Margins & Ad Efficiency<br><sup>Source: Empire Flippers P&L (Listing #92221)</sup>"),
                yaxis_title="%",
                height=350,
                margin=dict(l=60, r=20, t=50, b=40),
            )
            st.plotly_chart(fig, use_container_width=True)
            # AI analysis for margin chart
            margin_vals = [margin[m] for m in months_sorted]
            latest_margin = margin_vals[-1]
            tacos_latest = tacos.get(months_sorted[-1]) if tacos else None
            tacos_note = f" TaCOS is {tacos_latest:.1f}% — every 1% increase in ad spend reduces net margin by ~1 percentage point at current revenue." if tacos_latest else ""
            if latest_margin < 15:
                margin_stake = f"Net profit margin is {latest_margin:.1f}%. A 10% tariff increase or \\$0.50/unit cost rise would eliminate {min(latest_margin, 5):.0f}+ percentage points of margin at current pricing."
            else:
                margin_stake = f"Net profit margin is {latest_margin:.1f}%. The business retains {latest_margin - 10:.0f} percentage points of buffer above a 10% minimum viable margin."
            _ai_caption(f"{margin_stake}{tacos_note}")

        st.caption("Source: Empire Flippers verified P&L (Listing #92221)")


def _period_summary(products, label, cutoff_iso):
    """Render a summary block for a given time period cutoff."""
    st.markdown(f"**{label}**")
    rows = []
    for p in products:
        rh = [r for r in p["rating_history"] if r["date"][:10] >= cutoff_iso]
        ch = [c for c in p["review_count_history"] if c["date"][:10] >= cutoff_iso]
        if not rh and not ch:
            continue

        avg_r = sum(r["rating"] for r in rh) / len(rh) if rh else None
        reviews_added = (ch[-1]["count"] - ch[0]["count"]) if len(ch) >= 2 else None
        months_span = ((datetime.fromisoformat(ch[-1]["date"]) - datetime.fromisoformat(ch[0]["date"])).days / 30.44) if len(ch) >= 2 else None
        growth_rate = reviews_added / months_span if reviews_added is not None and months_span and months_span > 0 else None

        rows.append({
            "Product": product_name(p, short=True),
            "Avg Rating": f"{avg_r:.2f}" if avg_r else "\u2014",
            "Reviews Added": f"{reviews_added:,}" if reviews_added is not None else "\u2014",
            "Growth (rev/mo)": f"{growth_rate:.1f}" if growth_rate else "\u2014",
        })
    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


def page_yearly_breakdown(biz_data, biz_name):
    """Static yearly breakdown showing rating and review trends year over year.
    Static yearly breakdown showing rating and review trends year over year."""
    products = biz_data[biz_name]["products"]
    st.header(f"{biz_name} — Yearly Breakdown")
    st.caption("Source: Keepa API (keepa.com) | Full historical data")

    # Product overview table
    overview_rows = []
    for p in products:
        ch = p["review_count_history"]
        overview_rows.append({
            "Product": product_name(p, short=True),
            "ASIN": p["representative_asin"],
            "Reviews": f"{p['shared_reviews']:,}",
            "Variants": p["num_variations"],
            "Data Range": f"{ch[0]['date'][:7]} to {ch[-1]['date'][:7]}" if ch else "No data",
        })
    st.dataframe(pd.DataFrame(overview_rows), use_container_width=True, hide_index=True)

    # Filter to products with data
    products_with_data = [p for p in products if p["rating_history"] or p["review_count_history"]]
    if not products_with_data:
        st.info("No products with rating or review data.")
        return

    product_tabs = st.tabs([product_name(p, short=True) for p in products_with_data])

    for tab, p in zip(product_tabs, products_with_data):
      with tab:
        rh = p["rating_history"]
        ch = p["review_count_history"]
        if not rh and not ch:
            continue

        st.subheader(product_name(p, short=True))
        st.caption(
            f"ASIN: [{p['representative_asin']}]({amazon_link(p['representative_asin'])}) | "
            f"[Keepa]({keepa_link(p['representative_asin'])})"
        )
        if ch:
            range_text = (
                f"Review data: {ch[0]['date'][:7]} to {ch[-1]['date'][:7]} "
                f"({len(ch)} observations)"
            )
            if rh:
                range_text += (
                    f" | Rating data: {rh[0]['date'][:7]} to {rh[-1]['date'][:7]} "
                    f"({len(rh)} observations)"
                )
            else:
                range_text += " | No rating data"
            st.caption(range_text)

        # Group data by year
        rating_by_year = {}
        for r in rh:
            year = r["date"][:4]
            rating_by_year.setdefault(year, []).append(r["rating"])

        review_by_year = {}
        for c in ch:
            year = c["date"][:4]
            review_by_year.setdefault(year, []).append(c["count"])

        years = sorted(set(list(rating_by_year.keys()) + list(review_by_year.keys())))

        # Build year-over-year table
        rows = []
        prev_end_count = None
        last_known_avg = None
        last_known_min = None
        last_known_max = None
        for year in years:
            ratings = rating_by_year.get(year, [])
            counts = review_by_year.get(year, [])

            avg_rating = sum(ratings) / len(ratings) if ratings else None
            if avg_rating is not None:
                last_known_avg = avg_rating
                last_known_min = min(ratings)
                last_known_max = max(ratings)

            end_count = counts[-1] if counts else None
            start_count = counts[0] if counts else None
            reviews_added = (end_count - start_count) if end_count is not None and start_count is not None else None

            # Amazon removed reviews between end of prev year and start of this year
            if prev_end_count is not None and start_count is not None:
                gap = start_count - prev_end_count
                removed_str = f"{gap:,}" if gap < 0 else "0"
            else:
                removed_str = "No prior year"

            row = {
                "Year": year,
                "Avg Rating": f"{avg_rating:.2f}" if avg_rating else (f"~{last_known_avg:.2f} (carried forward)" if last_known_avg else "No data"),
                "Rating Min": f"{min(ratings):.1f}" if ratings else (f"~{last_known_min:.1f}" if last_known_min else "No data"),
                "Rating Max": f"{max(ratings):.1f}" if ratings else (f"~{last_known_max:.1f}" if last_known_max else "No data"),
                "Reviews (Start of Year)": f"{start_count:,}" if start_count is not None else "No data",
                "Reviews (End of Year)": f"{end_count:,}" if end_count is not None else "No data",
                "Net Reviews Added": f"{reviews_added:,}" if reviews_added is not None else "No data",
                "Reviews Removed Between Years": removed_str,
            }
            rows.append(row)
            prev_end_count = end_count

        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True)

        st.caption(
            '"Reviews Removed Between Years" = Reviews (Start of Year) minus Reviews (End of Previous Year). '
            "A negative number means Amazon removed reviews between those dates."
        )

        # Yearly charts
        chart_years = [row["Year"] for row in rows]

        # 1. Avg Rating by Year (bar chart)
        chart_ratings = []
        for row in rows:
            val = row["Avg Rating"]
            if val.startswith("~"):
                chart_ratings.append(float(val.split("~")[1].split(" ")[0]))
            elif val == "No data":
                chart_ratings.append(None)
            else:
                chart_ratings.append(float(val))

        if any(r is not None for r in chart_ratings):
            fig_r = go.Figure()
            fig_r.add_trace(go.Bar(
                x=chart_years,
                y=chart_ratings,
                marker_color=["#00d4aa" if r is not None and r >= 4.0 else "#ff6b6b" if r is not None else "#666" for r in chart_ratings],
                text=[f"{r:.2f}" if r is not None else "" for r in chart_ratings],
                textposition="outside",
            ))
            fig_r.add_hline(y=4.0, line_dash="dot", line_color="green", opacity=0.4)
            fig_r.update_layout(
                title=dict(text=f"Average Rating by Year — {product_name(p, short=True)}"),
                yaxis=dict(range=[2.5, 5.2], dtick=0.5, title="Avg Rating"),
                xaxis_title="Year",
                height=350,
                margin=dict(l=60, r=20, t=50, b=40),
                showlegend=False,
            )
            st.plotly_chart(fig_r, use_container_width=True)
            # AI analysis for yearly rating
            valid_ratings = [r for r in chart_ratings if r is not None]
            if len(valid_ratings) >= 2:
                first_yr_r, last_yr_r = valid_ratings[0], valid_ratings[-1]
                if abs(last_yr_r - first_yr_r) < 0.1:
                    _ai_caption(f"Rating consistent at ~{last_yr_r:.1f} across years. Stable ratings help keep customers buying — shoppers are more likely to trust a product with a steady track record.")
                elif last_yr_r > first_yr_r:
                    _ai_caption(f"Rating improved from {first_yr_r:.2f} to {last_yr_r:.2f} year-over-year. Higher ratings mean more customers choose this product over competitors, which typically leads to more sales.")
                else:
                    _ai_caption(f"Rating declined from {first_yr_r:.2f} to {last_yr_r:.2f} year-over-year. Lower ratings mean fewer customers choose this product — especially below 4.0, where shoppers tend to look elsewhere.")

        # 2. Net Reviews Added by Year (bar chart)
        chart_reviews = []
        for row in rows:
            val = row["Net Reviews Added"]
            if val == "No data":
                chart_reviews.append(None)
            else:
                chart_reviews.append(int(val.replace(",", "")))

        if any(r is not None for r in chart_reviews):
            fig_rev = go.Figure()
            fig_rev.add_trace(go.Bar(
                x=chart_years,
                y=[r if r is not None else 0 for r in chart_reviews],
                marker_color=["#3498db" if r is not None and r >= 0 else "#ff6b6b" if r is not None else "#666" for r in chart_reviews],
                text=[f"{r:,}" if r is not None else "" for r in chart_reviews],
                textposition="outside",
            ))
            fig_rev.update_layout(
                title=dict(text=f"Net Reviews Added by Year — {product_name(p, short=True)}"),
                yaxis_title="Reviews Added",
                xaxis_title="Year",
                height=350,
                margin=dict(l=60, r=20, t=50, b=40),
                showlegend=False,
            )
            st.plotly_chart(fig_rev, use_container_width=True)
            # AI analysis for yearly reviews
            valid_reviews = [r for r in chart_reviews if r is not None]
            if len(valid_reviews) >= 2:
                peak_yr_idx = valid_reviews.index(max(valid_reviews))
                peak_yr = chart_years[peak_yr_idx] if peak_yr_idx < len(chart_years) else "?"
                latest_rev = valid_reviews[-1]
                peak_rev = max(valid_reviews)
                if latest_rev < peak_rev * 0.5:
                    _ai_caption(f"Review acquisition peaked in {peak_yr} ({peak_rev:,} added) and has dropped to {latest_rev:,} — less than half the peak. Fewer new reviews usually means fewer sales, and competitors with more recent reviews will look more trustworthy to shoppers.")
                elif latest_rev >= peak_rev * 0.8:
                    _ai_caption(f"Review growth remains strong at {latest_rev:,}/year, near the peak of {peak_rev:,} in {peak_yr}. Consistent new reviews signal steady sales and help the listing stay trustworthy to new shoppers.")
                else:
                    _ai_caption(f"Review growth has slowed from a peak of {peak_rev:,} in {peak_yr} to {latest_rev:,} recently. Fewer new reviews makes it easier for competitors with faster review growth to look more trustworthy by comparison.")

        # Build carried-forward avg ratings for YoY comparison
        carried_avg_by_year = {}
        last_avg_for_yoy = None
        for year in years:
            ratings = rating_by_year.get(year, [])
            if ratings:
                last_avg_for_yoy = sum(ratings) / len(ratings)
            carried_avg_by_year[year] = last_avg_for_yoy

        # Year-over-year comparison
        if len(years) >= 2:
            st.markdown("**Year-over-Year Rating Trend** (avg rating change between years)")
            comparisons = []
            for i in range(1, len(years)):
                y1, y2 = years[i - 1], years[i]
                avg1 = carried_avg_by_year.get(y1)
                avg2 = carried_avg_by_year.get(y2)
                if avg1 is not None and avg2 is not None:
                    r1_actual = bool(rating_by_year.get(y1, []))
                    r2_actual = bool(rating_by_year.get(y2, []))
                    note = " (using carried-forward rating)" if not r1_actual or not r2_actual else ""
                    delta = avg2 - avg1
                    direction = "UP" if delta > 0.05 else "DOWN" if delta < -0.05 else "STABLE"
                    comparisons.append(
                        f"- {y1} to {y2}: avg rating {avg1:.2f} to {avg2:.2f} "
                        f"(rating change: {delta:+.2f}, trend: {direction}){note}"
                    )
                else:
                    missing = y1 if avg1 is None else y2
                    comparisons.append(f"- {y1} to {y2}: No rating data for {missing}")
            st.markdown("\n".join(comparisons))

        st.divider()

    # Period Summaries
    today = date.today()
    _period_summary(
        products,
        "Last 2 Years Summary",
        today.replace(year=today.year - 2).isoformat(),
    )
    st.markdown("")
    _period_summary(
        products,
        "Last 1 Year Summary",
        today.replace(year=today.year - 1).isoformat(),
    )
    st.markdown("")
    six_mo_year = today.year if today.month > 6 else today.year - 1
    six_mo_month = today.month - 6 if today.month > 6 else today.month + 6
    _period_summary(
        products,
        "Last 6 Months Summary",
        date(six_mo_year, six_mo_month, min(today.day, 28)).isoformat(),
    )


WALLAROO_CONVERSATION = """
**Dan Kim** — Thursday, 16 April 2:11 PM
> Accepted NDA and has been allowed access to view listing.
> Dan Kim is a premium buyer and pre-qualified by Flippa.

---

**Lazlo Cocheba** — Thursday, 16 April 2:12 PM
> Hi Dan,
>
> Thanks for signing the NDA. If you have any questions as you review the materials, feel free to reach out at any time. When you have a moment, it would be helpful to understand a bit about your professional background, how you typically fund acquisitions, and any initial questions you may have. If it's easier to discuss anything live, you can book time directly on my calendar here: https://meetings.hubspot.com/lazlo-cocheba
>
> Best,
> Lazlo

---

**Dan Kim** — Friday, 17 April 10:02 PM
> Hi Lazlo,
>
> Thank you for putting this listing together, the brand history and review base caught my attention, and I'd like to take a closer look before deciding whether to move forward.
>
> Would you be open to sharing a longer-term P&L, ideally going back to 2020 or earlier if available? I'd like to understand the full revenue and expense picture across those years.
>
> Looking forward to hearing from you.
>
> Best,
> Dan

---

**Lazlo Cocheba** — Saturday, 18 April 2:59 AM
> Hi Dan,
>
> We don't have a long-term P&L available at this stage, as that's typically shared post-LOI. That said, I've included a screenshot with high-level performance.
>
> Let me know if you have any questions. Happy to walk through the deal in more detail, feel free to grab time on my calendar.
>
> Best,
> Lazlo

---

**Dan Kim** — Tuesday, 21 April 5:16 PM
> Hi Lazlo,
>
> Thank you for sharing that. I'm interested in moving forward, if the deal comes in at or below a 24-month profit multiple, I'd be happy to submit a Letter of Intent.
> Could you confirm where things stand on that basis?
>
> Best,
> Dan

---

**Lazlo Cocheba** — Wednesday, 22 April 2:17 AM
> Hi Dan,
>
> We're open to exploring this. Could you submit an LOI with a preliminary offer? It would also be helpful to include a brief rationale for your valuation.
>
> If there's alignment, we're happy to move forward and provide any and all information requested.
>
> Best,
> Lazlo
"""


def page_seller_conversation():
    st.header("Wallaroo Wallets — Seller Conversation")
    st.caption("Ongoing conversation with Lazlo Cocheba via Flippa")
    st.markdown(
        "Listing: [Flippa](https://flippa.com/12260386-amazon-s-favorite-leather-phone-wallet-since-2016-"
        "hundreds-of-thousands-of-customers-over-7-000-reviews-for-the-hero-sku-with-an-average-4-6-star-rating)"
    )
    st.divider()
    st.markdown(WALLAROO_CONVERSATION)

    st.divider()
    st.markdown("#### Status")
    st.info(
        "Seller has asked Dan to submit an LOI with a preliminary offer and valuation rationale. "
        "P&L will be shared post-LOI. Dan's target: 24-month profit multiple or below."
    )


def page_methodology():
    st.header("Methodology & Definitions")
    st.caption("How metrics are calculated and data is sourced")

    st.markdown("""
### Data Source: Keepa API (keepa.com)

All data comes from [Keepa](https://keepa.com), which tracks Amazon product data continuously.
Keepa provides **complete, accurate** historical data for ratings and review counts.

- **Keepa Product API** with `rating=1` parameter returns csv indices 16 (rating) and 17 (review count)
- **Keepa Seller API** with `storefront=1` discovers all ASINs for a seller
- **Parent ASIN mapping** deduplicates product variations that share review pools

---

### Why We Track Historical Trends

A product with 4.6★ and 8,000 reviews looks great on the surface — but aggregate numbers can hide serious problems. This is why trend analysis is essential for acquisition due diligence:

#### 1. Aggregate ratings are sticky/lagging
A product with 8,000 reviews at 4.6★ could have quality problems for 6 months and the average barely moves. Recent trend catches this.

**Where this is answered in the dashboard:**
- **Executive Summary → Trend Table**: Rating column shows movement over 6mo/1yr/2yr (e.g. "4.6 → 4.5 (−0.10)")
- **Executive Summary → Trend Charts → Rating line chart**: Full time-series with the recent window highlighted in red — visual separation of "now" vs "then"
- **Executive Summary → Avg Rating bar chart**: Compares mean rating in recent vs prior period — directly catches recent quality decline even if the aggregate hasn't moved
- **[Business] page → Rating Trend KPI**: STABLE / IMPROVING / DECLINING label based on first-half vs second-half average

#### 2. Detects product/seller changes
Common Amazon pattern: product builds reputation, then seller cheapens materials or listing gets hijacked. Old reviews mask the change.

**Where this is partially answered:**
- **[Business] page → Notable Events → Rating Changes table**: Flags rating drops > 0.2 with "Possible quality issue or product change"
- **[Business] page → Notable Events → Price Changes table**: Shows price increases/decreases > 20% — price drops after a rating drop can signal material cheapening

**What we don't have:**
- No seller identity tracking — we can't detect if the seller behind the listing changed (listing hijack)
- No correlation between rating drops and price changes — you have to manually cross-reference the two tables
- No review text/sentiment analysis — we can't distinguish "quality got worse" from "shipping was slow" from a rating number alone

#### 3. Spots review manipulation
Sudden spikes in review volume (especially 5★ clusters) indicate paid reviews or giveaway campaigns. Healthy products accumulate reviews smoothly.

**Where this is answered:**
- **Executive Summary → Reviews/Mo Distribution box plot**: Shows whether review velocity is steady (tight box) or spiky (wide box with outliers) — spiky = suspicious
- **Executive Summary → Review Frequency line chart**: Visual time-series of reviews/month — organic growth looks smooth, manipulation shows as sharp spikes
- **[Business] page → Notable Events → Review Changes table**: Flags individual surge events ("Unusually high number of new reviews in a short period")
- **[Business] page → Review Integrity Assessment** (in expander): Quarterly growth consistency check — flags if any quarter's growth > 2× the average

**What we don't have:**
- No star-rating breakdown per period — we can't detect "5★ clusters" specifically, only total review volume spikes
- No review text analysis — can't check for generic/templated language typical of fake reviews
- No reviewer profile analysis — can't detect if reviews come from suspicious accounts

#### 4. Distinguishes "old hit" from "current hit"
Two products both at 4.5★/5,000 reviews look identical, but one might have gotten 90% of reviews 3 years ago and barely sells now.

**Where this is answered:**
- **Executive Summary → Trend Table → Reviews column**: Shows reviews added and per-month rate for 6mo/1yr/2yr — directly reveals if growth is current or historical
- **[Business] page → Review Growth KPI**: Reviews/month rate for the main product
- **[Business] page → Rating Trend KPI**: ACCELERATING / DECELERATING label
- **Executive Summary → Review Frequency line chart**: Visual time-series makes it immediately obvious if the product is still gaining reviews or flatlined years ago
- **[Evaluation] page → AI commentary**: Explicitly calls out when review acquisition has dropped below peak levels
- **[Yearly Breakdown] page**: Year-by-year review additions show exactly when growth happened

#### 5. Review removals are a signal
Amazon mass-purging reviews (visible as drops in review count charts) indicates past manipulation. This is valuable due diligence data.

**Where this is answered:**
- **[Business] page → Review Count chart**: Cumulative line chart — drops are visually obvious as downward steps. Caption: "Drops indicate reviews removed by Amazon"
- **[Business] page → Notable Events → Review Changes table**: Every purge event listed with date, size, and note "Amazon removed reviews — past review manipulation flagged"
- **[Business] page → Review Integrity Assessment**: Counts total purge events and total reviews removed; flags products with >100 removed
- **[Business] page → Price + Review overlay chart**: Green bars show monthly review additions with purge months excluded — lets you see the "clean" growth rate

**What we don't have:**
- No breakdown of purge severity — a single removal of 5 reviews vs. a mass purge of 680 reviews are both flagged, but the severity implications are very different
- No correlation with Amazon policy changes — some purges are platform-wide sweeps (less concerning) vs. targeted enforcement (more concerning)

---

### Key Metrics

#### Weighted Average Rating
$$
\\text{Weighted Avg Rating} = \\frac{\\sum_{i=1}^{n} (r_i \\times w_i)}{\\sum_{i=1}^{n} w_i}
$$

Where:
- $r_i$ = latest rating for product $i$
- $w_i$ = total review count for product $i$ (used as weight)
- $n$ = number of unique products

This gives more weight to products with more reviews, providing a fairer overall rating.

**Example:** Product A has 4.6 stars (1,000 reviews), Product B has 4.2 stars (200 reviews).
Weighted avg = (4.6 × 1,000 + 4.2 × 200) / (1,000 + 200) = 5,440 / 1,200 = **4.53**
Note: Product A's rating dominates because it has 5× more reviews — this prevents a low-volume product from unfairly dragging the average.

---

#### Review Growth Rate (per month)
$$
\\text{Growth/mo} = \\frac{C_{\\text{last}} - C_{\\text{first}}}{\\text{months}}
$$

Where:
- $C_{\\text{first}}$ = review count at first data point
- $C_{\\text{last}}$ = review count at last data point
- $\\text{months} = \\frac{\\text{days between first and last}}{30.44}$

**Example:** A product starts with 500 reviews in Jan 2022 and reaches 2,000 reviews by Jan 2024.
Growth = (2,000 - 500) / 24 months = **62.5 reviews/month**

---

#### Rating Trend Direction
Compares the average rating in the **first half** vs **second half** of the rating history:

$$
\\text{early\\_avg} = \\frac{1}{\\lfloor n/2 \\rfloor} \\sum_{i=1}^{\\lfloor n/2 \\rfloor} r_i \\qquad
\\text{recent\\_avg} = \\frac{1}{n - \\lfloor n/2 \\rfloor} \\sum_{i=\\lfloor n/2 \\rfloor+1}^{n} r_i
$$

$$
\\Delta = \\text{recent\\_avg} - \\text{early\\_avg}
$$

| Condition | Label |
|-----------|-------|
| $\\Delta > 0.05$ | IMPROVING |
| $\\Delta < -0.05$ | DECLINING |
| $|\\Delta| \\leq 0.05$ | STABLE |

**Example:** Rating history: [4.5, 4.4, 4.3, 4.6, 4.7, 4.8]
First half avg = (4.5 + 4.4 + 4.3) / 3 = 4.40
Second half avg = (4.6 + 4.7 + 4.8) / 3 = 4.70
Delta = +0.30 → **IMPROVING** (exceeds +0.05 threshold)

---

#### Growth Direction
Compares monthly review growth in the **first half** vs **second half** of the review count history:

$$
\\text{early\\_growth} = \\frac{C_{\\text{mid}} - C_{\\text{first}}}{\\text{months}_{\\text{first half}}} \\qquad
\\text{recent\\_growth} = \\frac{C_{\\text{last}} - C_{\\text{mid}}}{\\text{months}_{\\text{second half}}}
$$

| Condition | Label |
|-----------|-------|
| recent > early × 1.1 | ACCELERATING |
| recent < early × 0.9 | DECELERATING |
| otherwise | STEADY |

---

#### Parent ASIN Deduplication
Amazon product variations (colors, sizes) share a single review pool under a **parent ASIN**.
For example, Wallaroo's Phone Wallet has 14 color variants, all sharing ~7,107 reviews.

We use Keepa's `parentAsin` field to group child ASINs, then keep only one representative
per parent, avoiding double-counting reviews.

---

#### Sales Rank
Monthly average of Keepa's sales rank tracking (csv index 3). Lower rank = more sales.
The rank is category-specific (e.g., "Cell Phones & Accessories" for Wallaroo).

---

#### Review Count Drops
Review count charts show the raw cumulative total as reported by Keepa. Drops in the chart
indicate reviews removed by Amazon — this is normal Amazon behavior:
- **Small drops (1-9 reviews)**: Amazon removing individual policy-violating reviews
- **Large drops (100+)**: Amazon mass-purging reviews (e.g., suspected incentivized or fake reviews)

These drops are real data, not glitches, and represent valuable due diligence information.

---

### Combined Charts

**Combined Rating**: Weighted average of all products' ratings at each time point,
weighted by each product's review count at that time.

**Combined Reviews**: Sum of all products' review counts at each time point.
Missing data points are forward-filled before summing.

---

### Data Quality Notes
- Keepa data is sourced directly from Amazon's API — it represents the **true** rating and review count at each point in time
- Rating values from Keepa are integers (e.g., 46 = 4.6 stars), divided by 10 for display
- Keepa timestamps are in "Keepa minutes" (minutes since 2011-01-01 00:00 UTC)
- Review count drops in charts reflect real Amazon review removals, not data errors

---

### Limitations
- **No verified financials for Wallaroo** — Flippa listing does not include a P&L. Revenue/profit claims are unverified.
- **TeacherFav P&L is seller-reported** — Empire Flippers notes that product landed costs are submitted by the seller and not verified.
- **Review counts ≠ sales** — Review count correlates with but does not equal units sold. Typical review rates are 1-5% of purchases.
- **Sales rank is relative, not absolute** — Rank depends on category size and competition. A rank of 5,000 in "Cell Phones & Accessories" means different volume than 5,000 in "Toys & Games".
- **Rating data has gaps** — Keepa samples ratings less frequently than review counts. Some product/year combinations have no rating samples even though Keepa is still tracking review counts. The yearly breakdown carries forward the last known rating where possible (marked "carried forward" in the table).
- **No competitor analysis** — Dashboard only shows these products, not competing products in the same categories.
- **Data scope** — All pages use the full historical data from Keepa.

---

### Formula Verification
These formulas use standard e-commerce analytics approaches:
- **Weighted average** is the standard method for combining ratings across products of different sizes. An unweighted average would let a product with 16 reviews have equal influence to one with 7,107 reviews.
- **Linear growth rate** (reviews/month) provides a simple, interpretable measure. More complex models (exponential, polynomial) would overfit to noise in review data.
- **Half-split comparison** for trend direction is robust against outliers — comparing halves rather than individual data points avoids flagging temporary fluctuations as trends.
- **Review count as weight** for combined ratings reflects each product's market presence — a product customers actually buy should have proportionally more influence on the portfolio rating.
""")


# ─── Main App ───────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="Amazon FBA Review Analysis",
        page_icon="📊",
        layout="wide",
    )

    st.title("Amazon FBA Review Analysis Dashboard")
    st.caption("Wallaroo Wallets & TeacherFav | Keepa API Data | April 2026")

    biz_data = load_data()

    # Sidebar navigation
    page = st.sidebar.radio(
        "Navigation",
        [
            "Executive Summary",
            "Wallaroo Wallets",
            "TeacherFav",
            "Wallaroo — Evaluation",
            "TeacherFav — Evaluation",
            "Wallaroo — Yearly Breakdown",
            "TeacherFav — Yearly Breakdown",
            "Wallaroo — Seller Conversation",
            "Methodology & Definitions",
        ],
    )

    if page == "Executive Summary":
        page_executive_summary(biz_data)
    elif page == "Wallaroo Wallets":
        page_business_analysis(biz_data, "Wallaroo Wallets")
    elif page == "TeacherFav":
        page_business_analysis(biz_data, "TeacherFav")
    elif page == "Wallaroo — Evaluation":
        page_evaluation(biz_data, "Wallaroo Wallets")
    elif page == "TeacherFav — Evaluation":
        page_evaluation(biz_data, "TeacherFav")
    elif page == "Wallaroo — Yearly Breakdown":
        page_yearly_breakdown(biz_data, "Wallaroo Wallets")
    elif page == "TeacherFav — Yearly Breakdown":
        page_yearly_breakdown(biz_data, "TeacherFav")
    elif page == "Wallaroo — Seller Conversation":
        page_seller_conversation()
    elif page == "Methodology & Definitions":
        page_methodology()


if __name__ == "__main__":
    main()
